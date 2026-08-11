"""
FB Manager - Main Application
=============================
Flask app jo Facebook Pages, Team, aur Earnings manage karta hai

VERSION 2.0 - Phase 2+3+4 Update:
FB ACCOUNTS: Location, Issue Status, Profile Link + Username, DOB, Delete
PAGES: Recommendation (Okay/Not Okay), Fresh Start + Followers, Delete
BUSINESS MANAGERS: Partner Access tracking, Invite tracking, Delete
"""
import os
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, abort
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import func, extract, or_

from models import (
    db, User, FBAccount, BusinessManager, BMPartnerAccess,
    Page, Group, GroupAccount, BMInvite, DailyReport, TeamPayment, ActivityLog, init_db
)

# ==================== APP CONFIG ====================
app = Flask(__name__)
with app.app_context():
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-in-production-12345')

# Database - Railway pe PostgreSQL ya local SQLite
database_url = os.environ.get('DATABASE_URL', 'sqlite:///fb_manager.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==================== LOGIN MANAGER ====================
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Pehle login karein'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ==================== HELPERS ====================
def _parse_date(value):
    """Form se aayi date string ko date object banata hai"""
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _parse_int(value, default=0):
    """Form se aayi number string ko int banata hai"""
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


# ==================== ROLE DECORATORS ====================
def owner_required(f):
    """Sirf owner access kar sakta hai"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_owner():
            flash('Yeh page sirf owner ke liye hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def owner_or_supervisor(f):
    """Owner ya supervisor access kar sakta hai"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.is_worker():
            flash('Access denied', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def permission_required(perm):
    """
    Permission-based access. Owner ke paas hamesha sab kuch hai.
    Supervisor ka access owner ke diye hue flags par depend karta hai.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.has_perm(perm):
                return f(*args, **kwargs)
            flash('Iss kaam ki ijazat aap ke paas nahi hai. Owner se rabta karen.', 'error')
            return redirect(url_for('dashboard'))
        return decorated_function
    return decorator


# ==================== DATA SCOPING ====================
def _visible_user_ids():
    """
    Konse users ka data current user dekh sakta hai.
    Supervisor apni team ka data bhi dekhta hai.
    """
    ids = [current_user.id]
    if current_user.role == 'supervisor':
        ids += [w.id for w in User.query.filter_by(supervisor_id=current_user.id).all()]
    return ids


def _my_page_links():
    """
    Jo pages is user ko assign hain ya usne banaye —
    unki page IDs, FB account IDs aur BM IDs return karta hai.
    Isi se worker ko apne page ki FB ID / BM / Group bhi nazar aata hai.
    """
    ids = _visible_user_ids()
    rows = db.session.query(Page.id, Page.fb_account_id, Page.bm_id).filter(
        or_(Page.assigned_worker_id.in_(ids), Page.created_by_id.in_(ids))
    ).all()
    page_ids = {r[0] for r in rows if r[0]}
    account_ids = {r[1] for r in rows if r[1]}
    bm_ids = {r[2] for r in rows if r[2]}
    return page_ids, account_ids, bm_ids


def scope_records(query, model, assigned_field='assigned_user_id'):
    """
    Agar user ke paas 'view_all_data' nahi hai to sirf:
      - jo usne khud add kiya
      - jo uske (ya uski team ke) naam assign hua hai
      - AUR jo uske assigned pages se juda hua hai (FB ID / BM / Group)
    """
    if current_user.sees_all_data():
        return query

    ids = _visible_user_ids()
    conditions = []

    if hasattr(model, 'created_by_id'):
        conditions.append(model.created_by_id.in_(ids))
    if hasattr(model, assigned_field):
        conditions.append(getattr(model, assigned_field).in_(ids))

    # ---- Derived visibility: apne pages ke zariye ----
    if model is FBAccount:
        _pg, account_ids, _bm = _my_page_links()
        if account_ids:
            conditions.append(FBAccount.id.in_(account_ids))
    elif model is BusinessManager:
        _pg, _acc, bm_ids = _my_page_links()
        if bm_ids:
            conditions.append(BusinessManager.id.in_(bm_ids))
    elif model is Group:
        page_ids, account_ids, _bm = _my_page_links()
        if page_ids:
            conditions.append(Group.page_id.in_(page_ids))
        if account_ids:
            conditions.append(Group.fb_account_id.in_(account_ids))
            # extra IDs ke zariye bhi
            linked = db.session.query(GroupAccount.group_id).filter(
                GroupAccount.fb_account_id.in_(account_ids)).all()
            extra_group_ids = {g[0] for g in linked}
            if extra_group_ids:
                conditions.append(Group.id.in_(extra_group_ids))

    if not conditions:
        return query.filter(False)
    return query.filter(or_(*conditions))


def can_access(obj, assigned_field='assigned_user_id'):
    """Ek record par is user ka haq hai ya nahi (wahi logic jo list par lagta hai)"""
    if current_user.sees_all_data():
        return True
    model = type(obj)
    scoped = scope_records(model.query.filter(model.id == obj.id), model, assigned_field)
    return scoped.first() is not None


def deny_access():
    flash('Yeh record aap ko assign nahi hai. Owner se rabta karen.', 'error')
    return redirect(url_for('dashboard'))


def assignable_users():
    """Dropdown ke liye — team members jinko record assign kiya ja sakta hai"""
    return User.query.filter(User.role != 'owner', User.is_active == True).order_by(User.full_name).all()


# ==================== DUPLICATE DETECTION ====================
def _norm(value):
    """Comparison ke liye — lowercase, extra spaces hata kar"""
    if value is None:
        return ''
    return str(value).strip().lower()


# Facebook ke tracking parameters — inhen ignore karna hai
TRACKING_PARAMS = {
    'fbclid', 'ref', 'refid', 'refsrc', '__tn__', '__cft__', '__xts__', 'mibextid',
    'sfnsn', 'extid', 'locale', 'locale2', '_rdr', 'rdid', 'share_url', 'mid',
    'notif_id', 'notif_t', 'idorvanity', 'eav', 'paipv', 'source', 'utm_source',
    'utm_medium', 'utm_campaign', 'hoisted_section_header_type', 'comment_id',
}


def _norm_url(value):
    """
    URL normalize — protocol / www / m. / trailing slash hata kar.
    ZAROORI: profile.php?id=123 jaise URLs mein 'id' rakhna hai warna
    har profile URL doosre jaisa lagega. Sirf tracking params hatate hain.
    """
    v = _norm(value)
    if not v:
        return ''
    for p in ('https://', 'http://'):
        if v.startswith(p):
            v = v[len(p):]
    for p in ('www.', 'm.', 'mbasic.', 'web.', 'free.', 'touch.'):
        if v.startswith(p):
            v = v[len(p):]

    v = v.split('#')[0]

    if '?' in v:
        path, query = v.split('?', 1)
        kept = []
        for part in query.split('&'):
            if not part or '=' not in part:
                continue
            key, val = part.split('=', 1)
            key = key.strip()
            val = val.strip()
            if not key or not val or key in TRACKING_PARAMS:
                continue
            kept.append((key, val))
        kept.sort()
        path = path.rstrip('/')
        v = path + ('?' + '&'.join(f'{k}={val}' for k, val in kept) if kept else '')
    else:
        v = v.rstrip('/')
    return v


DUPLICATE_RULES = {
    'fb_account': {
        'model': lambda: FBAccount,
        'label': 'FB Account',
        'fields': [
            ('account_name', 'Account Name', 'text'),
            ('email', 'Email', 'text'),
            ('phone', 'Phone', 'text'),
            ('profile_username', 'Profile Username', 'text'),
            ('profile_link', 'Profile Link', 'url'),
        ],
        'name_field': 'account_name',
    },
    'page': {
        'model': lambda: Page,
        'label': 'Page',
        'fields': [
            ('page_name', 'Page Name', 'text'),
            ('page_url', 'Page URL', 'url'),
        ],
        'name_field': 'page_name',
    },
    'bm': {
        'model': lambda: BusinessManager,
        'label': 'Business Manager',
        'fields': [
            ('bm_name', 'BM Name', 'text'),
            ('bm_id', 'BM ID', 'text'),
        ],
        'name_field': 'bm_name',
    },
    'group': {
        'model': lambda: Group,
        'label': 'Group',
        'fields': [
            ('group_name', 'Group Name', 'text'),
            ('group_url', 'Group URL', 'url'),
            ('group_fb_id', 'Group FB ID', 'text'),
        ],
        'name_field': 'group_fb_id',
    },
}


def find_duplicate(kind, form_values, exclude_id=None):
    """
    Duplicate dhoondta hai — case-insensitive, URL normalize karke.
    Returns (field_label, existing_record) ya (None, None).
    """
    rule = DUPLICATE_RULES.get(kind)
    if not rule:
        return None, None

    model = rule['model']()
    records = model.query.all()

    for field, label, ftype in rule['fields']:
        entered = form_values.get(field)
        entered_n = _norm_url(entered) if ftype == 'url' else _norm(entered)
        if not entered_n:
            continue
        for rec in records:
            if exclude_id and rec.id == exclude_id:
                continue
            existing = getattr(rec, field, None)
            existing_n = _norm_url(existing) if ftype == 'url' else _norm(existing)
            if existing_n and existing_n == entered_n:
                return label, rec
    return None, None


def duplicate_message(kind, label, record):
    """Saaf message — kaun sa field, kis record se takra raha hai, kis ne add kiya"""
    rule = DUPLICATE_RULES.get(kind, {})
    entity = rule.get('label', 'Record')
    name_field = rule.get('name_field', 'id')
    existing_name = getattr(record, name_field, None) or f'#{record.id}'
    who = ''
    creator = getattr(record, 'created_by', None)
    if creator:
        who = f' — yeh {creator.full_name} ne add kiya tha'
    return (f'❌ Duplicate! Yeh {label} pehle se "{existing_name}" naam ke '
            f'{entity} mein mojood hai{who}. Dobara add nahi ho sakta.')


@app.route('/api/check-duplicate')
@login_required
def api_check_duplicate():
    """Live duplicate check — form mein type karte hi warning dikhane ke liye"""
    kind = request.args.get('type', '')
    field = request.args.get('field', '')
    value = request.args.get('value', '')
    exclude_id = request.args.get('exclude_id')

    rule = DUPLICATE_RULES.get(kind)
    if not rule or not value.strip():
        return jsonify({'duplicate': False})

    field_def = next((f for f in rule['fields'] if f[0] == field), None)
    if not field_def:
        return jsonify({'duplicate': False})

    try:
        exclude_id = int(exclude_id) if exclude_id else None
    except (TypeError, ValueError):
        exclude_id = None

    label, record = find_duplicate(kind, {field: value}, exclude_id=exclude_id)
    if record:
        name_field = rule.get('name_field', 'id')
        creator = getattr(record, 'created_by', None)
        return jsonify({
            'duplicate': True,
            'field_label': label,
            'existing_name': getattr(record, name_field, None) or f'#{record.id}',
            'added_by': creator.full_name if creator else None,
            'entity': rule['label'],
        })
    return jsonify({'duplicate': False})


# ==================== ACTIVITY LOG HELPER ====================
def log_activity(action, entity_type, entity_name, details=None):
    """Har create / update / delete ka record — khaas kar delete owner ko dikhane ke liye"""
    try:
        entry = ActivityLog(
            user_id=current_user.id if current_user.is_authenticated else None,
            user_name=current_user.full_name if current_user.is_authenticated else 'System',
            user_role=current_user.role if current_user.is_authenticated else '-',
            action=action,
            entity_type=entity_type,
            entity_name=(entity_name or '')[:180],
            details=details,
            # Owner apna kaam khud dekh chuka hai — sirf doosron ka unseen rahega
            seen_by_owner=bool(current_user.is_authenticated and current_user.is_owner())
        )
        db.session.add(entry)
        db.session.commit()
    except Exception:
        db.session.rollback()


# Templates mein permission check ke liye: {% if can('export') %}
@app.context_processor
def inject_permission_helper():
    def can(perm):
        if not current_user.is_authenticated:
            return False
        return current_user.has_perm(perm)
    return dict(can=can)


# ==================== AUTH ROUTES ====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Galat username ya password', 'error')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ==================== DASHBOARD ====================
@app.route('/')
@login_required
def dashboard():
    today = date.today()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)

    # Common stats (sab dekh sakte hain)
    total_pages = Page.query.filter_by(is_active=True).count()
    monetized_pages = Page.query.filter_by(is_active=True, monetization_status='monetized').count()
    non_monetized = Page.query.filter_by(is_active=True, monetization_status='non_monetized').count()
    in_review = Page.query.filter_by(is_active=True, monetization_status='in_review').count()
    total_fb_accounts = FBAccount.query.filter_by(status='active').count()

    # Role ke hisaab se data filter
    if current_user.is_worker():
        # Worker sirf apne assigned pages
        my_pages = Page.query.filter_by(assigned_worker_id=current_user.id, is_active=True).all()

        # Aaj ki report status
        today_reports = DailyReport.query.filter_by(
            worker_id=current_user.id,
            report_date=today
        ).count()

        # Last 7 days views (sirf views, NO earnings)
        my_views = db.session.query(func.sum(DailyReport.views)).filter(
            DailyReport.worker_id == current_user.id,
            DailyReport.report_date >= last_7_days
        ).scalar() or 0

        return render_template('dashboard_worker.html',
            my_pages=my_pages,
            total_pages_assigned=len(my_pages),
            today_reports=today_reports,
            pending_reports=len(my_pages) - today_reports,
            my_views=my_views
        )

    elif current_user.is_supervisor():
        # Supervisor: apni team ke pages
        team_workers = User.query.filter_by(supervisor_id=current_user.id).all()
        team_worker_ids = [w.id for w in team_workers]

        team_pages = Page.query.filter(
            Page.assigned_worker_id.in_(team_worker_ids),
            Page.is_active == True
        ).all() if team_worker_ids else []

        # Team views (NO earnings)
        team_views = db.session.query(func.sum(DailyReport.views)).filter(
            DailyReport.worker_id.in_(team_worker_ids),
            DailyReport.report_date >= last_7_days
        ).scalar() or 0 if team_worker_ids else 0

        return render_template('dashboard_supervisor.html',
            team_workers=team_workers,
            team_pages=team_pages,
            total_team_pages=len(team_pages),
            team_views=team_views,
            monetized_pages=monetized_pages,
            non_monetized=non_monetized
        )

    else:  # OWNER - full analytics dashboard
        # ---------- FB ACCOUNTS ----------
        all_accounts = FBAccount.query.all()
        total_accounts_all = len(all_accounts)
        active_accounts = sum(1 for a in all_accounts if a.status == 'active')
        accounts_with_issues = sum(1 for a in all_accounts if a.issue_type and a.issue_type != 'none')

        # By country
        country_map = {}
        for a in all_accounts:
            key = (a.location or '').strip() or 'Not Set'
            country_map[key] = country_map.get(key, 0) + 1
        country_data = sorted(country_map.items(), key=lambda x: x[1], reverse=True)

        # By issue type
        ISSUE_LABELS = {
            'none': 'No Issue',
            'whatsapp_code': 'WhatsApp Code',
            'unknown_number': 'Unknown Number',
            'banned': 'Banned',
            'password_issue': 'Password Issue',
            '2fa_issue': '2FA Issue',
            'other': 'Other',
        }
        issue_map = {}
        for a in all_accounts:
            key = a.issue_type or 'none'
            issue_map[key] = issue_map.get(key, 0) + 1
        issue_data = sorted(
            [(ISSUE_LABELS.get(k, k), v) for k, v in issue_map.items() if k != 'none'],
            key=lambda x: x[1], reverse=True
        )

        # By account status
        acc_status_map = {}
        for a in all_accounts:
            key = a.status or 'active'
            acc_status_map[key] = acc_status_map.get(key, 0) + 1

        # ---------- PAGES ----------
        all_pages = Page.query.filter_by(is_active=True).all()
        total_followers = sum(p.current_followers or 0 for p in all_pages)
        fresh_start_pages = sum(1 for p in all_pages if p.is_fresh_start)

        niche_map = {}
        for p in all_pages:
            key = (p.niche or '').strip() or 'Not Set'
            niche_map[key] = niche_map.get(key, 0) + 1
        niche_data = sorted(niche_map.items(), key=lambda x: x[1], reverse=True)

        reco_okay = sum(1 for p in all_pages if (p.recommendation or 'okay') == 'okay')
        reco_not_okay = len(all_pages) - reco_okay

        PSTATUS_LABELS = {
            'active': 'Active', 'suspended': 'Suspended', 'flagged': 'Flagged',
            'restricted': 'Restricted', 'unpublished': 'Unpublished',
            'under_review': 'Under Review', 'deleted': 'Deleted',
        }
        pstatus_map = {}
        for p in all_pages:
            key = p.page_status or 'active'
            pstatus_map[key] = pstatus_map.get(key, 0) + 1
        pstatus_data = [(PSTATUS_LABELS.get(k, k), v) for k, v in
                        sorted(pstatus_map.items(), key=lambda x: x[1], reverse=True)]
        pages_status_active = pstatus_map.get('active', 0)
        pages_problem = len(all_pages) - pages_status_active

        # Top pages by followers
        top_followers_pages = sorted(
            [p for p in all_pages if (p.current_followers or 0) > 0],
            key=lambda p: p.current_followers or 0, reverse=True
        )[:10]

        # ---------- GROUPS ----------
        all_groups = Group.query.all()
        total_groups = len(all_groups)
        total_group_members = sum(g.members_count or 0 for g in all_groups)
        active_groups = sum(1 for g in all_groups if (g.status or 'active') == 'active')
        groups_linked_page = sum(1 for g in all_groups if g.page_id)
        groups_linked_account = sum(1 for g in all_groups if g.fb_account_id and not g.page_id)
        top_groups = sorted(all_groups, key=lambda g: g.members_count or 0, reverse=True)[:5]

        # ---------- BUSINESS MANAGERS ----------
        all_bms = BusinessManager.query.all()
        total_bms = len(all_bms)
        active_bms = sum(1 for b in all_bms if b.status == 'active')
        total_partner_access = BMPartnerAccess.query.count()
        bms_with_invites = sum(1 for b in all_bms if b.invited_to_fb_account_id)
        bms_with_partners = sum(1 for b in all_bms if b.partner_accesses)

        # ---------- TEAM ----------
        team_members = User.query.filter(User.role != 'owner').all()
        total_team = len(team_members)
        total_workers = sum(1 for m in team_members if m.role == 'worker')
        total_supervisors = sum(1 for m in team_members if m.role == 'supervisor')

        # Team performance (last 7 days)
        team_perf = []
        for m in team_members:
            if not m.is_active:
                continue
            v = db.session.query(func.sum(DailyReport.views)).filter(
                DailyReport.worker_id == m.id,
                DailyReport.report_date >= last_7_days
            ).scalar() or 0
            assigned = Page.query.filter_by(assigned_worker_id=m.id, is_active=True).count()
            reports_count = DailyReport.query.filter(
                DailyReport.worker_id == m.id,
                DailyReport.report_date >= last_7_days
            ).count()
            team_perf.append({
                'name': m.full_name, 'role': m.role,
                'pages': assigned, 'views': int(v), 'reports': reports_count
            })
        team_perf = sorted(team_perf, key=lambda x: x['views'], reverse=True)[:8]

        # ---------- EARNINGS ----------
        total_earnings_30d = db.session.query(func.sum(DailyReport.earnings_usd)).filter(
            DailyReport.report_date >= last_30_days
        ).scalar() or 0

        total_earnings_today = db.session.query(func.sum(DailyReport.earnings_usd)).filter(
            DailyReport.report_date == today
        ).scalar() or 0

        total_views_30d = db.session.query(func.sum(DailyReport.views)).filter(
            DailyReport.report_date >= last_30_days
        ).scalar() or 0

        pending_fb_payments = db.session.query(func.sum(TeamPayment.total_earned_usd)).filter(
            TeamPayment.received_from_fb == False
        ).scalar() or 0

        pending_team_payments = db.session.query(func.sum(TeamPayment.agreed_amount_pkr)).filter(
            TeamPayment.received_from_fb == True,
            TeamPayment.paid_to_member == False
        ).scalar() or 0

        total_investment = sum(a.purchase_cost or 0 for a in all_accounts)

        top_pages = db.session.query(
            Page.page_name,
            func.sum(DailyReport.earnings_usd).label('earnings'),
            func.sum(DailyReport.views).label('views')
        ).join(DailyReport).filter(
            DailyReport.report_date >= last_30_days
        ).group_by(Page.id).order_by(func.sum(DailyReport.earnings_usd).desc()).limit(10).all()

        # Last 7 days chart data
        chart_data = []
        for i in range(7):
            d = today - timedelta(days=6-i)
            earnings = db.session.query(func.sum(DailyReport.earnings_usd)).filter(
                DailyReport.report_date == d
            ).scalar() or 0
            views = db.session.query(func.sum(DailyReport.views)).filter(
                DailyReport.report_date == d
            ).scalar() or 0
            chart_data.append({
                'date': d.strftime('%d %b'),
                'earnings': float(earnings),
                'views': int(views)
            })

        # ---------- HEALTH SCORE ----------
        signals = []
        if all_pages:
            signals.append(reco_okay / len(all_pages))
            signals.append(pages_status_active / len(all_pages))
        if total_accounts_all:
            signals.append((total_accounts_all - accounts_with_issues) / total_accounts_all)
            signals.append(active_accounts / total_accounts_all)
        if total_bms:
            signals.append(active_bms / total_bms)
        if total_groups:
            signals.append(active_groups / total_groups)
        health_score = int(round(sum(signals) / len(signals) * 100)) if signals else 100

        if health_score >= 90:
            health_label, health_color = 'Excellent', '#10b981'
        elif health_score >= 75:
            health_label, health_color = 'Good', '#42b72a'
        elif health_score >= 60:
            health_label, health_color = 'Fair', '#f59e0b'
        else:
            health_label, health_color = 'Needs Attention', '#dc2626'

        # ---------- DELETE ALERTS (owner ko pata chale) ----------
        recent_deletes = ActivityLog.query.filter_by(action='delete').order_by(
            ActivityLog.created_at.desc()).limit(6).all()
        unseen_deletes = ActivityLog.query.filter_by(action='delete', seen_by_owner=False).count()
        unseen_total = ActivityLog.query.filter_by(seen_by_owner=False).count()

        # ---------- RECENT ACTIVITY ----------
        recent_accounts = FBAccount.query.order_by(FBAccount.created_at.desc()).limit(4).all()
        recent_pages = Page.query.order_by(Page.created_at.desc()).limit(4).all()
        recent_groups = Group.query.order_by(Group.created_at.desc()).limit(3).all()

        # Attention list — jinko dekhna zaroori hai
        attention = []
        for a in all_accounts:
            if a.issue_type and a.issue_type != 'none':
                attention.append({'type': 'FB ID', 'name': a.account_name,
                                  'issue': ISSUE_LABELS.get(a.issue_type, a.issue_type),
                                  'url': url_for('edit_fb_account', account_id=a.id)})
        for p in all_pages:
            if (p.page_status or 'active') != 'active':
                attention.append({'type': 'Page', 'name': p.page_name,
                                  'issue': PSTATUS_LABELS.get(p.page_status, p.page_status),
                                  'url': url_for('edit_page', page_id=p.id)})
            elif (p.recommendation or 'okay') == 'not_okay':
                attention.append({'type': 'Page', 'name': p.page_name,
                                  'issue': 'Recommendation Not Okay',
                                  'url': url_for('edit_page', page_id=p.id)})
        for g in all_groups:
            if (g.status or 'active') != 'active':
                attention.append({'type': 'Group', 'name': g.group_name,
                                  'issue': (g.status or '').title(),
                                  'url': url_for('edit_group', group_id=g.id)})
        attention = attention[:12]

        return render_template('dashboard_owner.html',
            # counts
            total_pages=total_pages,
            monetized_pages=monetized_pages,
            non_monetized=non_monetized,
            in_review=in_review,
            total_fb_accounts=total_accounts_all,
            active_accounts=active_accounts,
            accounts_with_issues=accounts_with_issues,
            total_followers=total_followers,
            fresh_start_pages=fresh_start_pages,
            pages_status_active=pages_status_active,
            pages_problem=pages_problem,
            reco_okay=reco_okay,
            reco_not_okay=reco_not_okay,
            # groups
            total_groups=total_groups,
            total_group_members=total_group_members,
            active_groups=active_groups,
            groups_linked_page=groups_linked_page,
            groups_linked_account=groups_linked_account,
            top_groups=top_groups,
            # bms
            total_bms=total_bms,
            active_bms=active_bms,
            total_partner_access=total_partner_access,
            bms_with_invites=bms_with_invites,
            bms_with_partners=bms_with_partners,
            # team
            total_team=total_team,
            total_workers=total_workers,
            total_supervisors=total_supervisors,
            team_perf=team_perf,
            # money
            total_earnings_30d=total_earnings_30d,
            total_earnings_today=total_earnings_today,
            total_views_30d=total_views_30d,
            pending_fb_payments=pending_fb_payments,
            pending_team_payments=pending_team_payments,
            total_investment=total_investment,
            top_pages=top_pages,
            top_followers_pages=top_followers_pages,
            # charts
            chart_data=chart_data,
            country_data=country_data,
            niche_data=niche_data,
            issue_data=issue_data,
            pstatus_data=pstatus_data,
            acc_status_map=acc_status_map,
            # health + activity
            health_score=health_score,
            health_label=health_label,
            health_color=health_color,
            recent_accounts=recent_accounts,
            recent_pages=recent_pages,
            recent_groups=recent_groups,
            attention=attention,
            recent_deletes=recent_deletes,
            unseen_deletes=unseen_deletes,
            unseen_total=unseen_total
        )


# ==================== FB ACCOUNTS ====================
@app.route('/fb-accounts')
@login_required
@permission_required('view_fb_accounts')
def fb_accounts():
    query = scope_records(FBAccount.query, FBAccount)
    accounts = query.order_by(FBAccount.created_at.desc()).all()
    return render_template('fb_accounts.html', accounts=accounts)


@app.route('/fb-accounts/add', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def add_fb_account():
    if request.method == 'POST':
        # Duplicate check — same name/email/phone/username/link dobara na aaye
        dup_label, dup_rec = find_duplicate('fb_account', request.form)
        if dup_rec:
            flash(duplicate_message('fb_account', dup_label, dup_rec), 'error')
            return redirect(url_for('add_fb_account'))

        assigned = request.form.get('assigned_user_id')
        account = FBAccount(
            created_by_id=current_user.id,
            assigned_user_id=int(assigned) if assigned else None,
            account_name=request.form.get('account_name'),
            email=request.form.get('email'),
            phone=request.form.get('phone'),
            recovery_email=request.form.get('recovery_email'),
            purchase_date=_parse_date(request.form.get('purchase_date')),
            purchase_cost=float(request.form.get('purchase_cost') or 0),
            status=request.form.get('status', 'active'),
            notes=request.form.get('notes'),
            # ============ V2 FIELDS ============
            location=request.form.get('location') or None,
            issue_type=request.form.get('issue_type', 'none'),
            issue_notes=request.form.get('issue_notes') or None,
            profile_link=request.form.get('profile_link') or None,
            profile_username=request.form.get('profile_username') or None,
            date_of_birth=_parse_date(request.form.get('date_of_birth'))
        )
        # Encrypted fields
        account.set_fb_password(request.form.get('fb_password'))
        account.set_two_fa_code(request.form.get('two_fa_code'))

        db.session.add(account)
        db.session.commit()
        log_activity('create', 'FB Account', account.account_name)
        flash('FB Account add ho gaya (password encrypted hai)', 'success')
        return redirect(url_for('fb_accounts'))

    return render_template('fb_account_form.html', account=None, team_users=assignable_users())


@app.route('/fb-accounts/<int:account_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def edit_fb_account(account_id):
    account = FBAccount.query.get_or_404(account_id)
    if not can_access(account):
        return deny_access()
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('fb_account', request.form, exclude_id=account.id)
        if dup_rec:
            flash(duplicate_message('fb_account', dup_label, dup_rec), 'error')
            return redirect(url_for('edit_fb_account', account_id=account.id))
        if current_user.sees_all_data():
            assigned = request.form.get('assigned_user_id')
            account.assigned_user_id = int(assigned) if assigned else None
        account.account_name = request.form.get('account_name')
        account.email = request.form.get('email')
        account.phone = request.form.get('phone')

        # Password sirf tab update karen agar user ne diya hai
        new_password = request.form.get('fb_password')
        if new_password:
            account.set_fb_password(new_password)

        new_2fa = request.form.get('two_fa_code')
        if new_2fa:
            account.set_two_fa_code(new_2fa)

        account.recovery_email = request.form.get('recovery_email')
        account.purchase_date = _parse_date(request.form.get('purchase_date'))
        account.purchase_cost = float(request.form.get('purchase_cost') or 0)
        account.status = request.form.get('status', 'active')
        account.notes = request.form.get('notes')

        # ============ V2 FIELDS ============
        account.location = request.form.get('location') or None
        account.issue_type = request.form.get('issue_type', 'none')
        account.issue_notes = request.form.get('issue_notes') or None
        account.profile_link = request.form.get('profile_link') or None
        account.profile_username = request.form.get('profile_username') or None
        account.date_of_birth = _parse_date(request.form.get('date_of_birth'))

        db.session.commit()
        log_activity('update', 'FB Account', account.account_name)
        flash('FB Account update ho gaya', 'success')
        return redirect(url_for('fb_accounts'))

    return render_template('fb_account_form.html', account=account, team_users=assignable_users())


@app.route('/fb-accounts/<int:account_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_fb_account(account_id):
    """Delete an FB account (only if no pages/BMs linked)"""
    account = FBAccount.query.get_or_404(account_id)
    if not can_access(account):
        return deny_access()

    page_count = len(account.pages)
    bm_count = len(account.business_managers)

    if page_count > 0 or bm_count > 0:
        flash(
            f'Delete nahi ho sakta! Iss account ke sath {page_count} pages aur '
            f'{bm_count} BMs linked hain. Pehle unhen delete/reassign karen.',
            'error'
        )
        return redirect(url_for('fb_accounts'))

    account_name = account.account_name
    db.session.delete(account)
    db.session.commit()
    log_activity('delete', 'FB Account', account_name,
                 f'Email: {account.email or "-"} | Location: {account.location or "-"}')
    flash(f'FB Account "{account_name}" delete ho gaya', 'success')
    return redirect(url_for('fb_accounts'))


# ==================== FB ACCOUNTS - EXCEL IMPORT/EXPORT ====================
@app.route('/fb-accounts/export')
@login_required
@permission_required('export')
def export_fb_accounts():
    """Saare FB accounts ko Excel file mein export karen"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "FB Accounts"

    headers = [
        'Account Name', 'Email', 'Phone', 'FB Password',
        'Recovery Email', '2FA Code', 'Purchase Date',
        'Purchase Cost (PKR)', 'Status', 'Notes',
        'Location', 'Issue Type', 'Issue Notes', 'Profile Link', 'Username'
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1877F2', end_color='1877F2', fill_type='solid')
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    accounts = FBAccount.query.order_by(FBAccount.created_at.desc()).all()
    for acc in accounts:
        ws.append([
            acc.account_name or '',
            acc.email or '',
            acc.phone or '',
            acc.get_fb_password() or '',
            acc.recovery_email or '',
            acc.get_two_fa_code() or '',
            acc.purchase_date.strftime('%Y-%m-%d') if acc.purchase_date else '',
            acc.purchase_cost or 0,
            acc.status or 'active',
            acc.notes or '',
            acc.location or '',
            acc.issue_type or 'none',
            acc.issue_notes or '',
            acc.profile_link or '',
            acc.profile_username or ''
        ])

    column_widths = [20, 25, 15, 20, 25, 20, 14, 15, 12, 30, 15, 18, 30, 35, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fb_accounts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/fb-accounts/import', methods=['POST'])
@login_required
@permission_required('add_edit')
def import_fb_accounts():
    """Excel file se FB accounts ko bulk import karen"""
    from openpyxl import load_workbook

    if 'excel_file' not in request.files:
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('fb_accounts'))

    file = request.files['excel_file']
    if file.filename == '':
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('fb_accounts'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Sirf Excel file (.xlsx ya .xls) upload karen', 'error')
        return redirect(url_for('fb_accounts'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        imported_count = 0
        skipped_count = 0
        error_rows = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                account_name = str(row[0]).strip() if row[0] else None

                if not account_name:
                    skipped_count += 1
                    continue

                # V2 optional columns (11-15)
                def col(idx):
                    return str(row[idx]).strip() if len(row) > idx and row[idx] else None

                dup_vals = {'account_name': account_name, 'email': col(1), 'phone': col(2),
                            'profile_username': col(14), 'profile_link': col(13)}
                _lbl, existing = find_duplicate('fb_account', dup_vals)
                if existing:
                    skipped_count += 1
                    continue

                # Purchase date parse
                purchase_date = None
                if len(row) > 6 and row[6]:
                    try:
                        if isinstance(row[6], datetime):
                            purchase_date = row[6].date()
                        elif isinstance(row[6], date):
                            purchase_date = row[6]
                        else:
                            purchase_date = datetime.strptime(str(row[6]).split()[0], '%Y-%m-%d').date()
                    except Exception:
                        purchase_date = None

                try:
                    purchase_cost = float(row[7]) if len(row) > 7 and row[7] else 0
                except Exception:
                    purchase_cost = 0

                status = str(row[8]).strip().lower() if len(row) > 8 and row[8] else 'active'
                if status not in ['active', 'restricted', 'banned', 'disabled']:
                    status = 'active'

                issue_type = col(11) or 'none'
                valid_issues = ['none', 'whatsapp_code', 'unknown_number', 'banned',
                                'password_issue', '2fa_issue', 'other']
                if issue_type not in valid_issues:
                    issue_type = 'none'

                account = FBAccount(
                    created_by_id=current_user.id,
                    account_name=account_name,
                    email=col(1),
                    phone=col(2),
                    recovery_email=col(4),
                    purchase_date=purchase_date,
                    purchase_cost=purchase_cost,
                    status=status,
                    notes=col(9),
                    location=col(10),
                    issue_type=issue_type,
                    issue_notes=col(12),
                    profile_link=col(13),
                    profile_username=col(14)
                )
                if len(row) > 3 and row[3]:
                    account.set_fb_password(str(row[3]).strip())
                if len(row) > 5 and row[5]:
                    account.set_two_fa_code(str(row[5]).strip())

                db.session.add(account)
                imported_count += 1

            except Exception as e:
                error_rows.append(f"Row {row_num}: {str(e)}")
                continue

        db.session.commit()

        msg = f'{imported_count} accounts import ho gaye'
        if skipped_count > 0:
            msg += f' | {skipped_count} skip kiye gaye (already exist ya empty)'
        if error_rows:
            msg += f' | {len(error_rows)} errors'

        flash(msg, 'success' if imported_count > 0 else 'warning')

    except Exception as e:
        flash(f'Import error: {str(e)}', 'error')

    return redirect(url_for('fb_accounts'))


@app.route('/fb-accounts/template')
@login_required
@permission_required('export')
def fb_accounts_template():
    """Sample Excel template download karen for import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "FB Accounts Template"

    headers = [
        'Account Name', 'Email', 'Phone', 'FB Password',
        'Recovery Email', '2FA Code', 'Purchase Date',
        'Purchase Cost (PKR)', 'Status', 'Notes',
        'Location', 'Issue Type', 'Issue Notes', 'Profile Link', 'Username'
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1877F2', end_color='1877F2', fill_type='solid')
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    sample_rows = [
        ['Mudassar FB 1', 'fb1@gmail.com', '03001234567', 'YourPassword123',
         'recovery@gmail.com', '2FA backup code', '2024-01-15', 5000, 'active', 'Main account',
         'Pakistan', 'none', '', 'https://facebook.com/username1', 'username1'],
        ['USA FB 2', 'fb2@gmail.com', '03007654321', 'AnotherPass456',
         '', '', '2024-02-20', 3000, 'active', 'Secondary',
         'USA', 'whatsapp_code', 'Code WhatsApp par ja raha hai', '', ''],
    ]
    for row in sample_rows:
        ws.append(row)

    notes = [
        '',
        'Notes:',
        'Status: active, restricted, banned, disabled',
        'Issue Type: none, whatsapp_code, unknown_number, banned, password_issue, 2fa_issue, other',
        'Location: country ka naam (Pakistan, USA, India, Finland...)',
    ]
    for i, note in enumerate(notes, start=5):
        cell = ws.cell(row=i, column=1, value=note)
        if i > 5:
            cell.font = Font(italic=True, color='888888')

    column_widths = [20, 25, 15, 20, 25, 20, 14, 15, 12, 30, 15, 18, 30, 35, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='fb_accounts_template.xlsx'
    )


# ==================== API: Get decrypted password (owner only) ====================
@app.route('/api/fb-account/<int:account_id>/password')
@login_required
@permission_required('view_passwords')
def api_get_fb_password(account_id):
    """AJAX endpoint - returns decrypted password for display"""
    account = FBAccount.query.get_or_404(account_id)
    return jsonify({
        'success': True,
        'password': account.get_fb_password(),
        'two_fa': account.get_two_fa_code()
    })


# ==================== EXCEL HELPER ====================
def _excel_response(wb, filename):
    """Helper: workbook ko Excel download response banata hai"""
    from io import BytesIO
    from flask import send_file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _style_header(ws, headers, color='1877F2'):
    """Helper: header row par styling lagao"""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'


# ==================== PAGES - EXCEL IMPORT/EXPORT ====================
@app.route('/pages/export')
@login_required
@permission_required('export')
def export_pages():
    """Saare pages Excel mein export karen"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Pages"

    headers = [
        'Page Name', 'Page URL', 'Niche', 'FB Account', 'Business Manager',
        'Assigned Worker', 'Monetization Status', 'Monetized Date',
        'Page Created Date', 'Notes',
        'Recommendation', 'Recommendation Notes', 'Fresh Start',
        'Followers At Start', 'Current Followers', 'Page Status'
    ]
    _style_header(ws, headers, '42b72a')

    pages_list = Page.query.filter_by(is_active=True).all()
    for p in pages_list:
        ws.append([
            p.page_name or '',
            p.page_url or '',
            p.niche or '',
            p.fb_account.account_name if p.fb_account else '',
            p.business_manager.bm_name if p.business_manager else '',
            p.assigned_worker.full_name if p.assigned_worker else '',
            p.monetization_status or 'non_monetized',
            p.monetized_date.strftime('%Y-%m-%d') if p.monetized_date else '',
            p.page_created_date.strftime('%Y-%m-%d') if p.page_created_date else '',
            p.notes or '',
            p.recommendation or 'okay',
            p.recommendation_notes or '',
            'Yes' if p.is_fresh_start else 'No',
            p.followers_at_start or 0,
            p.current_followers or 0,
            p.page_status or 'active'
        ])

    column_widths = [25, 35, 15, 20, 20, 20, 18, 14, 14, 30, 15, 30, 12, 16, 16, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    filename = f"pages_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(wb, filename)


@app.route('/pages/template')
@login_required
@permission_required('export')
def pages_template():
    """Pages import ke liye sample Excel template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Pages Template"

    headers = [
        'Page Name', 'Page URL', 'Niche', 'FB Account', 'Business Manager',
        'Assigned Worker', 'Monetization Status', 'Monetized Date',
        'Page Created Date', 'Notes',
        'Recommendation', 'Recommendation Notes', 'Fresh Start',
        'Followers At Start', 'Current Followers', 'Page Status'
    ]
    _style_header(ws, headers, '42b72a')

    samples = [
        ['Comedy Hub', 'https://fb.com/comedyhub', 'Comedy', 'Mudassar FB 1', '',
         '', 'monetized', '2024-03-15', '2023-12-01', 'Top performer',
         'okay', '', 'Yes', 0, 45000, 'active'],
        ['News Pakistan', 'https://fb.com/newspk', 'News', 'Mudassar FB 1', 'Main BM',
         'Ali Worker', 'non_monetized', '', '2024-01-10', '',
         'not_okay', 'Recommendation issue', 'No', 12000, 15000, 'suspended'],
    ]
    for row in samples:
        ws.append(row)

    notes = [
        '',
        'Notes:',
        '1. FB Account: existing account_name use karen (exact match)',
        '2. Business Manager: optional, agar BM mein hai',
        '3. Assigned Worker: worker ka full_name (exact match)',
        '4. Monetization Status: monetized, non_monetized, in_review, suspended',
        '5. Dates: YYYY-MM-DD format mein (e.g. 2024-03-15)',
        '6. Recommendation: okay ya not_okay',
        '7. Fresh Start: Yes ya No (agar No to Followers At Start likhen)',
        '8. Page Status: active, suspended, flagged, restricted, unpublished'
    ]
    for i, note in enumerate(notes, start=5):
        cell = ws.cell(row=i, column=1, value=note)
        if i > 5:
            cell.font = Font(italic=True, color='888888', size=11)

    column_widths = [25, 35, 15, 20, 20, 20, 18, 14, 14, 30, 15, 30, 12, 16, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    return _excel_response(wb, 'pages_template.xlsx')


@app.route('/pages/import', methods=['POST'])
@login_required
@permission_required('add_edit')
def import_pages():
    """Excel se pages bulk import karen"""
    from openpyxl import load_workbook

    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('pages'))

    file = request.files['excel_file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Sirf Excel file upload karen', 'error')
        return redirect(url_for('pages'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                page_name = str(row[0]).strip() if row[0] else None
                if not page_name:
                    skipped += 1
                    continue

                _lbl, dup = find_duplicate('page', {
                    'page_name': page_name,
                    'page_url': str(row[1]).strip() if len(row) > 1 and row[1] else None})
                if dup:
                    skipped += 1
                    continue

                fb_account_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
                fb_account = FBAccount.query.filter_by(account_name=fb_account_name).first() if fb_account_name else None
                if not fb_account:
                    errors.append(f"Row {row_num}: FB Account '{fb_account_name}' nahi mila")
                    continue

                bm = None
                if len(row) > 4 and row[4]:
                    bm_name = str(row[4]).strip()
                    bm = BusinessManager.query.filter_by(bm_name=bm_name).first()

                worker = None
                if len(row) > 5 and row[5]:
                    worker_name = str(row[5]).strip()
                    worker = User.query.filter_by(full_name=worker_name, role='worker').first()

                def parse_date_cell(val):
                    if not val:
                        return None
                    if isinstance(val, datetime):
                        return val.date()
                    if isinstance(val, date):
                        return val
                    try:
                        return datetime.strptime(str(val).split()[0], '%Y-%m-%d').date()
                    except Exception:
                        return None

                monetization = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'non_monetized'
                if monetization not in ['monetized', 'non_monetized', 'in_review', 'suspended']:
                    monetization = 'non_monetized'

                # V2 columns
                recommendation = str(row[10]).strip().lower() if len(row) > 10 and row[10] else 'okay'
                if recommendation not in ['okay', 'not_okay']:
                    recommendation = 'okay'

                fresh_raw = str(row[12]).strip().lower() if len(row) > 12 and row[12] else 'yes'
                is_fresh = fresh_raw in ('yes', 'y', 'true', '1', 'haan')

                page = Page(
                    created_by_id=current_user.id,
                    page_name=page_name,
                    page_url=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    niche=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    fb_account_id=fb_account.id,
                    bm_id=bm.id if bm else None,
                    assigned_worker_id=worker.id if worker else None,
                    monetization_status=monetization,
                    monetized_date=parse_date_cell(row[7] if len(row) > 7 else None),
                    page_created_date=parse_date_cell(row[8] if len(row) > 8 else None),
                    notes=str(row[9]).strip() if len(row) > 9 and row[9] else None,
                    recommendation=recommendation,
                    recommendation_notes=str(row[11]).strip() if len(row) > 11 and row[11] else None,
                    is_fresh_start=is_fresh,
                    followers_at_start=_parse_int(row[13] if len(row) > 13 else 0),
                    current_followers=_parse_int(row[14] if len(row) > 14 else 0),
                    page_status=(str(row[15]).strip().lower()
                                 if len(row) > 15 and row[15]
                                 and str(row[15]).strip().lower() in
                                 ['active','suspended','flagged','restricted','unpublished']
                                 else 'active')
                )
                db.session.add(page)
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        msg = f'{imported} pages import ho gaye'
        if skipped > 0:
            msg += f' | {skipped} skip kiye'
        if errors:
            msg += f' | {len(errors)} errors: ' + '; '.join(errors[:3])
        flash(msg, 'success' if imported > 0 else 'warning')
    except Exception as e:
        flash(f'Import error: {str(e)}', 'error')

    return redirect(url_for('pages'))


# ==================== BUSINESS MANAGERS ====================
@app.route('/business-managers')
@login_required
@permission_required('view_bms')
def business_managers():
    """BM list page"""
    bms = scope_records(BusinessManager.query, BusinessManager).order_by(
        BusinessManager.created_at.desc()).all()
    return render_template('business_managers.html', bms=bms)


@app.route('/business-managers/add', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def add_business_manager():
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('bm', request.form)
        if dup_rec:
            flash(duplicate_message('bm', dup_label, dup_rec), 'error')
            return redirect(url_for('add_business_manager'))

        invited_id = request.form.get('invited_to_fb_account_id')
        assigned = request.form.get('assigned_user_id')

        bm = BusinessManager(
            created_by_id=current_user.id,
            assigned_user_id=int(assigned) if assigned else None,
            bm_name=request.form.get('bm_name'),
            bm_id=request.form.get('bm_id'),
            fb_account_id=int(request.form.get('fb_account_id')),
            status=request.form.get('status', 'active'),
            # ============ V2 FIELDS ============
            invited_to_fb_account_id=int(invited_id) if invited_id else None,
            invite_date=_parse_date(request.form.get('invite_date')),
            invite_notes=request.form.get('invite_notes') or None
        )
        db.session.add(bm)
        db.session.commit()

        # Add form par likhe hue email invites bhi save karen
        invite_count = _save_invite_emails(bm, request.form)

        log_activity('create', 'Business Manager', bm.bm_name,
                     f'{invite_count} email invite(s)' if invite_count else None)
        msg = 'Business Manager add ho gaya.'
        if invite_count:
            msg += f' {invite_count} email invite bhi save ho gaye.'
        msg += ' Ab Partner Access add kar sakte hain.'
        flash(msg, 'success')
        return redirect(url_for('edit_business_manager', bm_id=bm.id))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    return render_template('bm_form.html', bm=None, fb_accounts=fb_accounts_list, partners=[],
                           invites=[], team_users=assignable_users())


@app.route('/business-managers/<int:bm_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def edit_business_manager(bm_id):
    bm = BusinessManager.query.get_or_404(bm_id)
    if not can_access(bm):
        return deny_access()
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('bm', request.form, exclude_id=bm.id)
        if dup_rec:
            flash(duplicate_message('bm', dup_label, dup_rec), 'error')
            return redirect(url_for('edit_business_manager', bm_id=bm.id))
        if current_user.sees_all_data():
            assigned_u = request.form.get('assigned_user_id')
            bm.assigned_user_id = int(assigned_u) if assigned_u else None
        invited_id = request.form.get('invited_to_fb_account_id')

        bm.bm_name = request.form.get('bm_name')
        bm.bm_id = request.form.get('bm_id')
        bm.fb_account_id = int(request.form.get('fb_account_id'))
        bm.status = request.form.get('status', 'active')

        # ============ V2 FIELDS ============
        bm.invited_to_fb_account_id = int(invited_id) if invited_id else None
        bm.invite_date = _parse_date(request.form.get('invite_date'))
        bm.invite_notes = request.form.get('invite_notes') or None

        db.session.commit()
        log_activity('update', 'Business Manager', bm.bm_name)
        flash('Business Manager update ho gaya', 'success')
        return redirect(url_for('business_managers'))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    partners = BMPartnerAccess.query.filter_by(source_bm_id=bm.id).order_by(
        BMPartnerAccess.created_at.desc()
    ).all()
    invites = BMInvite.query.filter_by(bm_id=bm.id).order_by(BMInvite.created_at.desc()).all()
    return render_template('bm_form.html', bm=bm, fb_accounts=fb_accounts_list, partners=partners,
                           invites=invites, team_users=assignable_users())


@app.route('/business-managers/<int:bm_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_business_manager(bm_id):
    """Delete a BM (only if no pages linked)"""
    bm = BusinessManager.query.get_or_404(bm_id)
    if not can_access(bm):
        return deny_access()

    page_count = Page.query.filter_by(bm_id=bm.id).count()
    if page_count > 0:
        flash(
            f'Delete nahi ho sakta! Iss BM ke sath {page_count} pages linked hain. '
            f'Pehle unhen delete/reassign karen.',
            'error'
        )
        return redirect(url_for('business_managers'))

    bm_name = bm.bm_name
    partner_n = len(bm.partner_accesses)
    db.session.delete(bm)  # partner_accesses cascade delete ho jayen ge
    db.session.commit()
    log_activity('delete', 'Business Manager', bm_name,
                 f'{partner_n} partner access records bhi delete hue')
    flash(f'Business Manager "{bm_name}" delete ho gaya', 'success')
    return redirect(url_for('business_managers'))


# ==================== BM PARTNER ACCESS (V2 NEW) ====================
@app.route('/business-managers/<int:bm_id>/partners/add', methods=['POST'])
@login_required
@permission_required('add_edit')
def add_bm_partner(bm_id):
    """Kisi BM ko doosri BM ka partner access diya - record karen"""
    bm = BusinessManager.query.get_or_404(bm_id)

    partner_name = request.form.get('partner_bm_name')
    if not partner_name:
        flash('Partner BM ka naam zaroori hai', 'error')
        return redirect(url_for('edit_business_manager', bm_id=bm.id))

    partner = BMPartnerAccess(
        source_bm_id=bm.id,
        partner_bm_name=partner_name,
        partner_bm_id=request.form.get('partner_bm_id') or None,
        access_granted_date=_parse_date(request.form.get('access_granted_date')),
        access_level=request.form.get('access_level') or None,
        notes=request.form.get('partner_notes') or None,
        is_active=True
    )
    db.session.add(partner)
    db.session.commit()
    log_activity('create', 'Partner Access', partner_name, f'Source BM: {bm.bm_name}')
    flash(f'Partner access record add ho gaya: {partner_name}', 'success')
    return redirect(url_for('edit_business_manager', bm_id=bm.id))


@app.route('/business-managers/<int:bm_id>/partners/<int:partner_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_bm_partner(bm_id, partner_id):
    """Partner access record delete karen"""
    partner = BMPartnerAccess.query.get_or_404(partner_id)
    if partner.source_bm_id != bm_id:
        flash('Galat request', 'error')
        return redirect(url_for('business_managers'))

    partner_name = partner.partner_bm_name
    source_name = partner.source_bm.bm_name if partner.source_bm else '-'
    db.session.delete(partner)
    db.session.commit()
    log_activity('delete', 'Partner Access', partner_name, f'Source BM: {source_name}')
    flash(f'Partner access "{partner_name}" delete ho gaya', 'success')
    return redirect(url_for('edit_business_manager', bm_id=bm_id))


def _save_invite_emails(bm, form, default_status='pending'):
    """
    Form ke 'invite_emails' field se emails nikal kar BM ke sath save karta hai.
    Comma / nayi line / semicolon / space — kisi se bhi alag kiye ja sakte hain.
    Duplicate emails skip ho jate hain. Kitne add hue woh return karta hai.
    """
    emails_raw = form.get('invite_emails') or ''
    parts = [e.strip() for e in
             emails_raw.replace('\n', ',').replace(';', ',').replace(' ', ',').split(',')]
    emails = [e for e in parts if e and '@' in e]
    if not emails:
        return 0

    invited_date = _parse_date(form.get('invite_sent_date'))
    notes = form.get('invite_email_notes') or None
    acc_id = form.get('invite_fb_account_id')
    status = form.get('invite_status', default_status)
    if status not in ('pending', 'accepted', 'expired'):
        status = default_status

    added = 0
    for email in emails:
        exists = BMInvite.query.filter(
            BMInvite.bm_id == bm.id,
            func.lower(BMInvite.email) == email.lower()
        ).first()
        if exists:
            continue
        db.session.add(BMInvite(
            bm_id=bm.id,
            email=email,
            status=status,
            invited_date=invited_date,
            accepted_date=date.today() if status == 'accepted' else None,
            fb_account_id=int(acc_id) if acc_id else None,
            notes=notes
        ))
        added += 1

    if added:
        db.session.commit()
    return added


# ==================== BM EMAIL INVITES (pending / accepted) ====================
@app.route('/business-managers/<int:bm_id>/invites/add', methods=['POST'])
@login_required
@permission_required('add_edit')
def add_bm_invite(bm_id):
    """BM ka invite kis email par bheja — jitne chahen email add karen"""
    bm = BusinessManager.query.get_or_404(bm_id)
    if not can_access(bm):
        return deny_access()

    raw = request.form.get('invite_emails') or ''
    if '@' not in raw:
        flash('Kam az kam ek sahi email likhen', 'error')
        return redirect(url_for('edit_business_manager', bm_id=bm.id))

    added = _save_invite_emails(bm, request.form)

    if added:
        log_activity('create', 'BM Invite', bm.bm_name, f'{added} email invite(s)')
        flash(f'{added} email invite add ho gaye', 'success')
    else:
        flash('Yeh email(s) pehle se mojood hain', 'warning')
    return redirect(url_for('edit_business_manager', bm_id=bm.id))


@app.route('/business-managers/<int:bm_id>/invites/<int:invite_id>/status', methods=['POST'])
@login_required
@permission_required('add_edit')
def update_bm_invite_status(bm_id, invite_id):
    """Invite ka status badlen — pending / accepted / expired"""
    invite = BMInvite.query.get_or_404(invite_id)
    if invite.bm_id != bm_id:
        flash('Galat request', 'error')
        return redirect(url_for('business_managers'))

    new_status = request.form.get('status', 'pending')
    if new_status not in ('pending', 'accepted', 'expired'):
        new_status = 'pending'

    invite.status = new_status
    invite.accepted_date = date.today() if new_status == 'accepted' else None

    acc_id = request.form.get('fb_account_id')
    if acc_id:
        invite.fb_account_id = int(acc_id)

    db.session.commit()
    log_activity('update', 'BM Invite', invite.email, f'Status: {new_status}')
    flash(f'Invite status update ho gaya: {new_status}', 'success')
    return redirect(url_for('edit_business_manager', bm_id=bm_id))


@app.route('/business-managers/<int:bm_id>/invites/<int:invite_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_bm_invite(bm_id, invite_id):
    invite = BMInvite.query.get_or_404(invite_id)
    if invite.bm_id != bm_id:
        flash('Galat request', 'error')
        return redirect(url_for('business_managers'))

    email = invite.email
    db.session.delete(invite)
    db.session.commit()
    log_activity('delete', 'BM Invite', email)
    flash(f'Invite "{email}" delete ho gaya', 'success')
    return redirect(url_for('edit_business_manager', bm_id=bm_id))


# ==================== BUSINESS MANAGERS - EXCEL ====================
@app.route('/business-managers/export')
@login_required
@permission_required('export')
def export_business_managers():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Managers"

    headers = [
        'BM Name', 'BM ID Number', 'Linked FB Account', 'Status', 'Pages Count',
        'Invited To FB Account', 'Invite Date', 'Invite Notes', 'Partner Access Count',
        'Pending Invite Emails', 'Accepted Invite Emails'
    ]
    _style_header(ws, headers, '7F77DD')

    bms = BusinessManager.query.all()
    for bm in bms:
        pages_count = Page.query.filter_by(bm_id=bm.id, is_active=True).count()
        partner_count = BMPartnerAccess.query.filter_by(source_bm_id=bm.id).count()
        pending = [i.email for i in bm.email_invites if (i.status or 'pending') == 'pending']
        accepted = [i.email for i in bm.email_invites if i.status == 'accepted']
        ws.append([
            bm.bm_name,
            bm.bm_id or '',
            bm.fb_account.account_name if bm.fb_account else '',
            bm.status,
            pages_count,
            bm.invited_fb_account.account_name if bm.invited_fb_account else '',
            bm.invite_date.strftime('%Y-%m-%d') if bm.invite_date else '',
            bm.invite_notes or '',
            partner_count,
            ', '.join(pending),
            ', '.join(accepted)
        ])

    for i, w in enumerate([25, 25, 25, 15, 15, 25, 14, 30, 18, 40, 40], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, f"business_managers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/business-managers/partners/export')
@login_required
@permission_required('export')
def export_bm_partners():
    """Saare partner access records export karen"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BM Partner Access"

    headers = [
        'Source BM Name', 'Source BM ID', 'Partner BM Name', 'Partner BM ID',
        'Access Level', 'Access Granted Date', 'Active', 'Notes'
    ]
    _style_header(ws, headers, '7F77DD')

    partners = BMPartnerAccess.query.all()
    for p in partners:
        source = p.source_bm
        ws.append([
            source.bm_name if source else '',
            source.bm_id if source else '',
            p.partner_bm_name,
            p.partner_bm_id or '',
            p.access_level or '',
            p.access_granted_date.strftime('%Y-%m-%d') if p.access_granted_date else '',
            'Yes' if p.is_active else 'No',
            p.notes or ''
        ])

    for i, w in enumerate([25, 22, 25, 22, 20, 18, 10, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, f"bm_partner_access_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/business-managers/template')
@login_required
@permission_required('export')
def bm_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    headers = ['BM Name', 'BM ID Number', 'Linked FB Account', 'Status']
    _style_header(ws, headers, '7F77DD')
    ws.append(['Main BM', '123456789012345', 'Mudassar FB 1', 'active'])
    ws.append(['Backup BM', '987654321098765', 'Mudassar FB 2', 'active'])
    ws.cell(row=5, column=1, value='Note: FB Account exact match honi chahiye (account_name)').font = Font(italic=True, color='888888')
    ws.cell(row=6, column=1, value='Note: Partner Access aur Invite tracking BM edit page se add karen').font = Font(italic=True, color='888888')
    for i, w in enumerate([25, 25, 25, 15], 1):
        ws.column_dimensions[chr(64+i)].width = w
    return _excel_response(wb, 'business_managers_template.xlsx')


@app.route('/business-managers/import', methods=['POST'])
@login_required
@permission_required('add_edit')
def import_business_managers():
    from openpyxl import load_workbook
    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('business_managers'))

    try:
        wb = load_workbook(request.files['excel_file'], data_only=True)
        ws = wb.active
        imported, skipped, errors = 0, 0, []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            try:
                bm_name = str(row[0]).strip() if row[0] else None
                if not bm_name:
                    skipped += 1
                    continue
                _lbl, dup = find_duplicate('bm', {
                    'bm_name': bm_name,
                    'bm_id': str(row[1]).strip() if len(row) > 1 and row[1] else None})
                if dup:
                    skipped += 1
                    continue

                fb_acc_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                fb_acc = FBAccount.query.filter_by(account_name=fb_acc_name).first() if fb_acc_name else None
                if not fb_acc:
                    errors.append(f"Row {row_num}: FB Account nahi mila")
                    continue

                bm = BusinessManager(
                    created_by_id=current_user.id,
                    bm_name=bm_name,
                    bm_id=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    fb_account_id=fb_acc.id,
                    status=str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'active'
                )
                db.session.add(bm)
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        msg = f'{imported} BMs import | {skipped} skip'
        if errors:
            msg += f' | {len(errors)} errors'
        flash(msg, 'success' if imported > 0 else 'warning')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('business_managers'))


# ==================== TEAM - EXCEL ====================
@app.route('/team/export')
@login_required
@permission_required('export')
def export_team():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Members"

    headers = ['Full Name', 'Username', 'Role', 'Phone', 'Supervisor', 'Active', 'Joined']
    _style_header(ws, headers, 'd62976')

    members = User.query.filter(User.role != 'owner').all()
    for m in members:
        ws.append([
            m.full_name,
            m.username,
            m.role,
            m.phone or '',
            m.supervisor.full_name if m.supervisor else '',
            'Yes' if m.is_active else 'No',
            m.created_at.strftime('%Y-%m-%d') if m.created_at else ''
        ])

    for i, w in enumerate([25, 18, 15, 15, 25, 10, 14], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, f"team_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/team/template')
@login_required
@permission_required('export')
def team_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    headers = ['Full Name', 'Username', 'Password', 'Role', 'Phone', 'Supervisor Name']
    _style_header(ws, headers, 'd62976')
    ws.append(['Ali Khan', 'ali_khan', 'password123', 'worker', '03001234567', 'Ahmad Supervisor'])
    ws.append(['Ahmad Hussain', 'ahmad_h', 'mypass456', 'supervisor', '03007654321', ''])

    notes = [
        '',
        'Notes:',
        '1. Role: worker, supervisor (owner allowed nahi)',
        '2. Supervisor Name: workers ke liye apne supervisor ka full_name',
        '3. Username unique hona chahiye',
        '4. Password kam az kam 6 characters'
    ]
    for i, note in enumerate(notes, start=5):
        cell = ws.cell(row=i, column=1, value=note)
        if i > 5:
            cell.font = Font(italic=True, color='888888', size=11)

    for i, w in enumerate([25, 18, 18, 15, 15, 25], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, 'team_template.xlsx')


@app.route('/team/import', methods=['POST'])
@login_required
@permission_required('add_edit')
def import_team():
    from openpyxl import load_workbook
    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('team'))

    try:
        wb = load_workbook(request.files['excel_file'], data_only=True)
        ws = wb.active
        imported, skipped, errors = 0, 0, []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            try:
                full_name = str(row[0]).strip() if row[0] else None
                username = str(row[1]).strip() if len(row) > 1 and row[1] else None
                password = str(row[2]).strip() if len(row) > 2 and row[2] else None
                role = str(row[3]).strip().lower() if len(row) > 3 and row[3] else None

                if not (full_name and username and password and role):
                    skipped += 1
                    continue

                if role not in ['worker', 'supervisor']:
                    errors.append(f"Row {row_num}: Invalid role '{role}'")
                    continue

                if User.query.filter_by(username=username).first():
                    skipped += 1
                    continue

                supervisor = None
                if role == 'worker' and len(row) > 5 and row[5]:
                    supervisor = User.query.filter_by(
                        full_name=str(row[5]).strip(), role='supervisor'
                    ).first()

                user = User(
                    full_name=full_name,
                    username=username,
                    role=role,
                    phone=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                    supervisor_id=supervisor.id if supervisor else None,
                    is_active=True
                )
                user.set_password(password)
                db.session.add(user)
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        msg = f'{imported} members import | {skipped} skip'
        if errors:
            msg += f' | {len(errors)} errors: ' + '; '.join(errors[:3])
        flash(msg, 'success' if imported > 0 else 'warning')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('team'))


# ==================== PAGES ====================
@app.route('/pages')
@login_required
def pages():
    # Role-based filtering
    if current_user.is_worker():
        # Worker: apne assign kiye + apne banaye hue pages (dono)
        pages_list = Page.query.filter(
            or_(Page.assigned_worker_id == current_user.id,
                Page.created_by_id == current_user.id),
            Page.is_active == True
        ).order_by(Page.created_at.desc()).all()
    elif current_user.is_supervisor() and not current_user.sees_all_data():
        team_worker_ids = [w.id for w in User.query.filter_by(supervisor_id=current_user.id).all()]
        visible_ids = team_worker_ids + [current_user.id]
        pages_list = Page.query.filter(
            or_(Page.assigned_worker_id.in_(visible_ids),
                Page.created_by_id.in_(visible_ids)),
            Page.is_active == True
        ).order_by(Page.created_at.desc()).all()
    else:  # owner ya supervisor jise sab dikhta hai
        pages_list = scope_records(
            Page.query.filter_by(is_active=True), Page, assigned_field='assigned_worker_id'
        ).order_by(Page.created_at.desc()).all()

    return render_template('pages.html', pages=pages_list)


@app.route('/pages/add', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def add_page():
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('page', request.form)
        if dup_rec:
            flash(duplicate_message('page', dup_label, dup_rec), 'error')
            return redirect(url_for('add_page'))

        is_fresh = request.form.get('is_fresh_start') == 'yes'

        page = Page(
            created_by_id=current_user.id,
            page_name=request.form.get('page_name'),
            page_url=request.form.get('page_url'),
            niche=request.form.get('niche'),
            fb_account_id=int(request.form.get('fb_account_id')),
            bm_id=int(request.form.get('bm_id')) if request.form.get('bm_id') else None,
            assigned_worker_id=int(request.form.get('assigned_worker_id')) if request.form.get('assigned_worker_id') else None,
            monetization_status=request.form.get('monetization_status', 'non_monetized'),
            page_created_date=_parse_date(request.form.get('page_created_date')),
            monetized_date=_parse_date(request.form.get('monetized_date')),
            notes=request.form.get('notes'),
            # ============ V2 FIELDS ============
            recommendation=request.form.get('recommendation', 'okay'),
            recommendation_notes=request.form.get('recommendation_notes') or None,
            is_fresh_start=is_fresh,
            followers_at_start=0 if is_fresh else _parse_int(request.form.get('followers_at_start')),
            current_followers=_parse_int(request.form.get('current_followers')),
            page_status=request.form.get('page_status', 'active'),
            page_status_notes=request.form.get('page_status_notes') or None
        )
        db.session.add(page)
        db.session.commit()
        log_activity('create', 'Page', page.page_name)
        flash('Page add ho gaya', 'success')
        return redirect(url_for('pages'))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    workers = User.query.filter(User.role.in_(['worker', 'supervisor']), User.is_active == True).all()
    bms = scope_records(BusinessManager.query.filter_by(status='active'), BusinessManager).all()
    return render_template('page_form.html', page=None, fb_accounts=fb_accounts_list, workers=workers, bms=bms)


@app.route('/pages/<int:page_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def edit_page(page_id):
    page = Page.query.get_or_404(page_id)
    if not can_access(page, assigned_field='assigned_worker_id'):
        return deny_access()
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('page', request.form, exclude_id=page.id)
        if dup_rec:
            flash(duplicate_message('page', dup_label, dup_rec), 'error')
            return redirect(url_for('edit_page', page_id=page.id))
        is_fresh = request.form.get('is_fresh_start') == 'yes'

        page.page_name = request.form.get('page_name')
        page.page_url = request.form.get('page_url')
        page.niche = request.form.get('niche')
        page.fb_account_id = int(request.form.get('fb_account_id'))
        page.bm_id = int(request.form.get('bm_id')) if request.form.get('bm_id') else None
        page.assigned_worker_id = int(request.form.get('assigned_worker_id')) if request.form.get('assigned_worker_id') else None
        page.monetization_status = request.form.get('monetization_status', 'non_monetized')
        page.page_created_date = _parse_date(request.form.get('page_created_date'))
        page.monetized_date = _parse_date(request.form.get('monetized_date'))
        page.notes = request.form.get('notes')

        # ============ V2 FIELDS ============
        page.recommendation = request.form.get('recommendation', 'okay')
        page.recommendation_notes = request.form.get('recommendation_notes') or None
        page.is_fresh_start = is_fresh
        page.followers_at_start = 0 if is_fresh else _parse_int(request.form.get('followers_at_start'))
        page.current_followers = _parse_int(request.form.get('current_followers'))
        page.page_status = request.form.get('page_status', 'active')
        page.page_status_notes = request.form.get('page_status_notes') or None

        db.session.commit()
        log_activity('update', 'Page', page.page_name)
        flash('Page update ho gaya', 'success')
        return redirect(url_for('pages'))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    workers = User.query.filter(User.role.in_(['worker', 'supervisor']), User.is_active == True).all()
    bms = scope_records(BusinessManager.query.filter_by(status='active'), BusinessManager).all()
    return render_template('page_form.html', page=page, fb_accounts=fb_accounts_list, workers=workers, bms=bms)


@app.route('/pages/<int:page_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_page(page_id):
    """Delete a page (reports bhi delete ho jayen ge)"""
    page = Page.query.get_or_404(page_id)
    if not can_access(page, assigned_field='assigned_worker_id'):
        return deny_access()

    report_count = DailyReport.query.filter_by(page_id=page.id).count()
    page_name = page.page_name

    # Reports delete karen pehle (foreign key constraint)
    if report_count > 0:
        DailyReport.query.filter_by(page_id=page.id).delete()

    db.session.delete(page)
    db.session.commit()
    log_activity('delete', 'Page', page_name,
                 f'{report_count} reports bhi delete hue')

    msg = f'Page "{page_name}" delete ho gaya'
    if report_count > 0:
        msg += f' (uske {report_count} reports bhi delete hue)'
    flash(msg, 'success')
    return redirect(url_for('pages'))


def _save_group_extra_accounts(group, form):
    """
    Group ke extra FB IDs (checkbox list) save karta hai.
    Main ID alag hai — usay repeat nahi karte.
    """
    selected = set()
    for raw in form.getlist('extra_account_ids'):
        try:
            val = int(raw)
        except (TypeError, ValueError):
            continue
        if group.fb_account_id and val == group.fb_account_id:
            continue  # main ID dobara nahi
        selected.add(val)

    existing = {ga.fb_account_id: ga for ga in
                GroupAccount.query.filter_by(group_id=group.id).all()}

    # jo hata diye gaye
    for acc_id, ga in existing.items():
        if acc_id not in selected:
            db.session.delete(ga)

    # naye add karen
    for acc_id in selected:
        if acc_id not in existing:
            db.session.add(GroupAccount(group_id=group.id, fb_account_id=acc_id))

    db.session.commit()


# ==================== GROUPS (V2 NEW - Feature 7) ====================
@app.route('/groups')
@login_required
@permission_required('view_groups')
def groups():
    """Saare groups ki list"""
    groups_list = scope_records(Group.query, Group).order_by(Group.created_at.desc()).all()
    return render_template('groups.html', groups=groups_list)


@app.route('/groups/add', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def add_group():
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('group', request.form)
        if dup_rec:
            flash(duplicate_message('group', dup_label, dup_rec), 'error')
            return redirect(url_for('add_group'))

        page_id = request.form.get('page_id')
        fb_account_id = request.form.get('fb_account_id')
        assigned = request.form.get('assigned_user_id')

        group = Group(
            created_by_id=current_user.id,
            assigned_user_id=int(assigned) if assigned else None,
            group_name=request.form.get('group_name'),
            group_url=request.form.get('group_url') or None,
            group_fb_id=request.form.get('group_fb_id') or None,
            page_id=int(page_id) if page_id else None,
            fb_account_id=int(fb_account_id) if fb_account_id else None,
            members_count=_parse_int(request.form.get('members_count')),
            status=request.form.get('status', 'active'),
            notes=request.form.get('notes') or None
        )
        db.session.add(group)
        db.session.commit()

        # Extra IDs (jin mein yeh group bhi admin hai)
        _save_group_extra_accounts(group, request.form)

        log_activity('create', 'Group', group.group_name,
                     f'{len(group.all_account_ids)} FB ID(s) linked')
        flash('Group add ho gaya', 'success')
        return redirect(url_for('groups'))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    pages_list = scope_records(Page.query.filter_by(is_active=True), Page,
                               assigned_field='assigned_worker_id').order_by(Page.page_name).all()
    return render_template('group_form.html', group=None, fb_accounts=fb_accounts_list,
                           pages=pages_list, team_users=assignable_users(),
                           extra_selected=set())


@app.route('/groups/<int:group_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('add_edit')
def edit_group(group_id):
    group = Group.query.get_or_404(group_id)
    if not can_access(group):
        return deny_access()
    if request.method == 'POST':
        dup_label, dup_rec = find_duplicate('group', request.form, exclude_id=group.id)
        if dup_rec:
            flash(duplicate_message('group', dup_label, dup_rec), 'error')
            return redirect(url_for('edit_group', group_id=group.id))
        if current_user.sees_all_data():
            assigned_u = request.form.get('assigned_user_id')
            group.assigned_user_id = int(assigned_u) if assigned_u else None
        page_id = request.form.get('page_id')
        fb_account_id = request.form.get('fb_account_id')

        group.group_name = request.form.get('group_name')
        group.group_url = request.form.get('group_url') or None
        group.group_fb_id = request.form.get('group_fb_id') or None
        group.page_id = int(page_id) if page_id else None
        group.fb_account_id = int(fb_account_id) if fb_account_id else None
        group.members_count = _parse_int(request.form.get('members_count'))
        group.status = request.form.get('status', 'active')
        group.notes = request.form.get('notes') or None

        _save_group_extra_accounts(group, request.form)

        db.session.commit()
        log_activity('update', 'Group', group.group_name,
                     f'{len(group.all_account_ids)} FB ID(s) linked')
        flash('Group update ho gaya', 'success')
        return redirect(url_for('groups'))

    fb_accounts_list = scope_records(FBAccount.query.filter_by(status='active'), FBAccount).all()
    pages_list = scope_records(Page.query.filter_by(is_active=True), Page,
                               assigned_field='assigned_worker_id').order_by(Page.page_name).all()
    extra_selected = {ga.fb_account_id for ga in group.extra_accounts}
    return render_template('group_form.html', group=group, fb_accounts=fb_accounts_list,
                           pages=pages_list, team_users=assignable_users(),
                           extra_selected=extra_selected)


@app.route('/groups/<int:group_id>/delete', methods=['POST'])
@login_required
@permission_required('delete')
def delete_group(group_id):
    group = Group.query.get_or_404(group_id)
    if not can_access(group):
        return deny_access()
    group_name = group.group_name
    members = group.members_count or 0
    db.session.delete(group)
    db.session.commit()
    log_activity('delete', 'Group', group_name, f'{members:,} members')
    flash(f'Group "{group_name}" delete ho gaya', 'success')
    return redirect(url_for('groups'))


@app.route('/groups/export')
@login_required
@permission_required('export')
def export_groups():
    """Saare groups Excel mein export karen"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Groups"

    headers = [
        'Group Name', 'Group URL', 'Group FB ID', 'Linked Page',
        'Main FB Account', 'Extra FB IDs (admin)', 'Total IDs', 'Members', 'Status', 'Notes'
    ]
    _style_header(ws, headers, '0ea5e9')

    for g in Group.query.all():
        extras = ', '.join(ga.fb_account.account_name for ga in g.extra_accounts if ga.fb_account)
        ws.append([
            g.group_name or '',
            g.group_url or '',
            g.group_fb_id or '',
            g.page.page_name if g.page else '',
            g.fb_account.account_name if g.fb_account else '',
            extras,
            len(g.all_account_ids),
            g.members_count or 0,
            g.status or 'active',
            g.notes or ''
        ])

    for i, w in enumerate([28, 35, 20, 25, 22, 40, 12, 14, 14, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, f"groups_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/groups/template')
@login_required
@permission_required('export')
def groups_template():
    """Groups import ke liye sample template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Groups Template"

    headers = [
        'Group Name', 'Group URL', 'Group FB ID', 'Linked Page',
        'Linked FB Account', 'Members', 'Status', 'Notes'
    ]
    _style_header(ws, headers, '0ea5e9')

    ws.append(['Comedy Lovers PK', 'https://facebook.com/groups/comedypk', '123456789',
               'Comedy Hub', '', 45000, 'active', 'Main group'])
    ws.append(['News Discussion', 'https://facebook.com/groups/newspk', '',
               '', 'Mudassar FB 1', 12000, 'active', ''])

    notes = [
        '',
        'Notes:',
        '1. Linked Page: page ka exact naam (optional)',
        '2. Linked FB Account: account ka exact naam (optional)',
        '3. Group ya to page se linked ho sakta hai ya FB account se (ya dono khali)',
        '4. Status: active, restricted, suspended, deleted',
    ]
    for i, note in enumerate(notes, start=5):
        cell = ws.cell(row=i, column=1, value=note)
        if i > 5:
            cell.font = Font(italic=True, color='888888', size=11)

    for i, w in enumerate([28, 35, 20, 25, 22, 14, 14, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    return _excel_response(wb, 'groups_template.xlsx')


@app.route('/groups/import', methods=['POST'])
@login_required
@permission_required('add_edit')
def import_groups():
    """Excel se groups bulk import"""
    from openpyxl import load_workbook

    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('Koi file select nahi ki gayi', 'error')
        return redirect(url_for('groups'))

    try:
        wb = load_workbook(request.files['excel_file'], data_only=True)
        ws = wb.active
        imported, skipped, errors = 0, 0, []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            try:
                group_name = str(row[0]).strip() if row[0] else None
                if not group_name:
                    skipped += 1
                    continue
                _lbl, dup = find_duplicate('group', {
                    'group_name': group_name,
                    'group_url': str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    'group_fb_id': str(row[2]).strip() if len(row) > 2 and row[2] else None})
                if dup:
                    skipped += 1
                    continue

                page_obj = None
                if len(row) > 3 and row[3]:
                    page_obj = Page.query.filter_by(page_name=str(row[3]).strip()).first()

                acc_obj = None
                if len(row) > 4 and row[4]:
                    acc_obj = FBAccount.query.filter_by(account_name=str(row[4]).strip()).first()

                status = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'active'
                if status not in ['active', 'restricted', 'suspended', 'deleted']:
                    status = 'active'

                group = Group(
                    created_by_id=current_user.id,
                    group_name=group_name,
                    group_url=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    group_fb_id=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    page_id=page_obj.id if page_obj else None,
                    fb_account_id=acc_obj.id if acc_obj else None,
                    members_count=_parse_int(row[5] if len(row) > 5 else 0),
                    status=status,
                    notes=str(row[7]).strip() if len(row) > 7 and row[7] else None
                )
                db.session.add(group)
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        msg = f'{imported} groups import ho gaye'
        if skipped > 0:
            msg += f' | {skipped} skip kiye'
        if errors:
            msg += f' | {len(errors)} errors'
        flash(msg, 'success' if imported > 0 else 'warning')
    except Exception as e:
        flash(f'Import error: {str(e)}', 'error')

    return redirect(url_for('groups'))


# ==================== DAILY REPORTS ====================
@app.route('/reports/submit', methods=['GET', 'POST'])
@login_required
def submit_report():
    today = date.today()

    if current_user.is_worker():
        my_pages = Page.query.filter_by(assigned_worker_id=current_user.id, is_active=True).all()
    elif current_user.is_supervisor():
        team_worker_ids = [w.id for w in User.query.filter_by(supervisor_id=current_user.id).all()]
        my_pages = Page.query.filter(Page.assigned_worker_id.in_(team_worker_ids + [current_user.id])).all()
    else:
        my_pages = Page.query.filter_by(is_active=True).all()

    if request.method == 'POST':
        page_id = int(request.form.get('page_id'))
        report_date_str = request.form.get('report_date', today.strftime('%Y-%m-%d'))
        report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()

        # Check duplicate
        existing = DailyReport.query.filter_by(page_id=page_id, report_date=report_date).first()
        if existing:
            existing.views = _parse_int(request.form.get('views'))
            existing.reach = _parse_int(request.form.get('reach'))
            existing.followers_gained = _parse_int(request.form.get('followers_gained'))
            if current_user.is_owner():
                existing.earnings_usd = float(request.form.get('earnings_usd') or 0)
            existing.notes = request.form.get('notes')
            flash('Report update ho gayi', 'success')
        else:
            report = DailyReport(
                page_id=page_id,
                worker_id=current_user.id,
                report_date=report_date,
                views=_parse_int(request.form.get('views')),
                reach=_parse_int(request.form.get('reach')),
                followers_gained=_parse_int(request.form.get('followers_gained')),
                earnings_usd=float(request.form.get('earnings_usd') or 0) if current_user.is_owner() else 0,
                notes=request.form.get('notes')
            )
            db.session.add(report)
            flash('Report submit ho gayi', 'success')

        db.session.commit()
        return redirect(url_for('submit_report'))

    today_reports = DailyReport.query.filter_by(report_date=today).all()
    submitted_page_ids = [r.page_id for r in today_reports if r.worker_id == current_user.id]

    return render_template('submit_report.html',
        pages=my_pages,
        today=today,
        submitted_page_ids=submitted_page_ids
    )


@app.route('/reports')
@login_required
def reports_list():
    """Reports list with role-based earnings visibility"""
    days = int(request.args.get('days', 7))
    start_date = date.today() - timedelta(days=days)

    query = DailyReport.query.filter(DailyReport.report_date >= start_date)

    if current_user.is_worker():
        query = query.filter(DailyReport.worker_id == current_user.id)
    elif current_user.is_supervisor():
        team_ids = [w.id for w in User.query.filter_by(supervisor_id=current_user.id).all()]
        team_ids.append(current_user.id)
        query = query.filter(DailyReport.worker_id.in_(team_ids))

    reports = query.order_by(DailyReport.report_date.desc()).all()
    return render_template('reports_list.html', reports=reports, days=days)


# ==================== TEAM MANAGEMENT ====================
@app.route('/team')
@login_required
@permission_required('view_team')
def team():
    if current_user.is_owner():
        members = User.query.filter(User.role != 'owner').order_by(User.role, User.full_name).all()
    else:
        members = User.query.filter_by(supervisor_id=current_user.id).all()
    return render_template('team.html', members=members)


@app.route('/team/add', methods=['GET', 'POST'])
@login_required
@owner_required
def add_team_member():
    if request.method == 'POST':
        user = User(
            username=request.form.get('username'),
            full_name=request.form.get('full_name'),
            role=request.form.get('role'),
            phone=request.form.get('phone'),
            supervisor_id=int(request.form.get('supervisor_id')) if request.form.get('supervisor_id') else None
        )
        # Naye member ko by default kuch nazar nahi aata — owner permissions deta hai
        if user.role == 'worker':
            for f in ['can_view_fb_accounts', 'can_view_passwords', 'can_view_bms',
                      'can_view_groups', 'can_view_team', 'can_add_edit', 'can_delete',
                      'can_export', 'can_view_all_data']:
                setattr(user, f, False)
        else:
            user.can_view_passwords = False
            user.can_delete = False
            user.can_export = False
            user.can_view_all_data = False
        user.perms_initialized = True
        user.set_password(request.form.get('password'))
        db.session.add(user)
        db.session.commit()
        log_activity('create', 'Team Member', user.full_name, f'Role: {user.role}')
        flash(f'{user.full_name} add ho gaya — ab Team page se permissions set karen', 'success')
        return redirect(url_for('team'))

    supervisors = User.query.filter_by(role='supervisor', is_active=True).all()
    return render_template('team_form.html', member=None, supervisors=supervisors)


# ==================== PAYMENTS (OWNER ONLY) ====================
@app.route('/payments')
@login_required
@owner_required
def payments():
    payments_list = TeamPayment.query.order_by(TeamPayment.month.desc(), TeamPayment.created_at.desc()).all()

    total_pending_fb = sum(p.total_earned_usd for p in payments_list if not p.received_from_fb)
    total_pending_pay = sum(p.agreed_amount_pkr for p in payments_list if p.received_from_fb and not p.paid_to_member)
    total_paid = sum(p.agreed_amount_pkr for p in payments_list if p.paid_to_member)

    return render_template('payments.html',
        payments=payments_list,
        total_pending_fb=total_pending_fb,
        total_pending_pay=total_pending_pay,
        total_paid=total_paid
    )


@app.route('/payments/add', methods=['GET', 'POST'])
@login_required
@owner_required
def add_payment():
    if request.method == 'POST':
        payment = TeamPayment(
            user_id=int(request.form.get('user_id')),
            month=request.form.get('month'),
            total_earned_usd=float(request.form.get('total_earned_usd') or 0),
            agreed_amount_pkr=float(request.form.get('agreed_amount_pkr') or 0),
            notes=request.form.get('notes')
        )
        db.session.add(payment)
        db.session.commit()
        flash('Payment entry add ho gayi', 'success')
        return redirect(url_for('payments'))

    members = User.query.filter(User.role.in_(['supervisor', 'worker']), User.is_active == True).all()
    return render_template('payment_form.html', members=members)


@app.route('/payments/<int:pid>/mark-received', methods=['POST'])
@login_required
@owner_required
def mark_received(pid):
    payment = TeamPayment.query.get_or_404(pid)
    payment.received_from_fb = True
    payment.received_date = date.today()
    db.session.commit()
    flash('FB se received mark ho gaya', 'success')
    return redirect(url_for('payments'))


@app.route('/payments/<int:pid>/mark-paid', methods=['POST'])
@login_required
@owner_required
def mark_paid(pid):
    payment = TeamPayment.query.get_or_404(pid)
    payment.paid_to_member = True
    payment.payment_date = date.today()
    payment.payment_method = request.form.get('payment_method', 'Cash')
    payment.payment_reference = request.form.get('payment_reference')
    db.session.commit()
    flash('Payment paid mark ho gaya', 'success')
    return redirect(url_for('payments'))



# ==================== ACCOUNT SETTINGS — PASSWORD CHANGE ====================
@app.route('/settings/password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Har user apna password khud change kar sakta hai (hashed store hota hai)"""
    if request.method == 'POST':
        current_pw = request.form.get('current_password') or ''
        new_pw = request.form.get('new_password') or ''
        confirm_pw = request.form.get('confirm_password') or ''

        if not current_user.check_password(current_pw):
            flash('Purana password galat hai', 'error')
            return redirect(url_for('change_password'))

        if len(new_pw) < 6:
            flash('Naya password kam az kam 6 characters ka hona chahiye', 'error')
            return redirect(url_for('change_password'))

        if new_pw != confirm_pw:
            flash('Naya password aur confirm password match nahi kar rahe', 'error')
            return redirect(url_for('change_password'))

        if new_pw == current_pw:
            flash('Naya password purane se alag hona chahiye', 'error')
            return redirect(url_for('change_password'))

        current_user.set_password(new_pw)
        db.session.commit()
        log_activity('update', 'Password', current_user.full_name, 'Apna password change kiya')
        flash('Password change ho gaya. Agli baar naya password use karen.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')


# ==================== TEAM — EDIT + PERMISSIONS + RESET PASSWORD ====================
PERMISSION_FIELDS = [
    ('can_view_fb_accounts', '🔑 FB Accounts dekh sakta hai'),
    ('can_view_passwords', '👁️ FB passwords dekh sakta hai (sensitive)'),
    ('can_view_bms', '🏢 Business Managers dekh sakta hai'),
    ('can_view_groups', '👥 Groups dekh sakta hai'),
    ('can_view_team', '🧑‍💼 Team list dekh sakta hai'),
    ('can_add_edit', '➕ Naya add / edit kar sakta hai (import bhi)'),
    ('can_delete', '🗑️ Delete kar sakta hai'),
    ('can_export', '📤 Excel download / export kar sakta hai'),
    ('can_view_all_data', '🌐 SAB ka data dekh sakta hai (off = sirf apna + assigned)'),
]


@app.route('/team/<int:member_id>/edit', methods=['GET', 'POST'])
@login_required
@owner_required
def edit_team_member(member_id):
    """Owner team member ki details aur permissions set karta hai"""
    member = User.query.get_or_404(member_id)
    if member.role == 'owner':
        flash('Owner account yahan se edit nahi hota', 'error')
        return redirect(url_for('team'))

    if request.method == 'POST':
        member.full_name = request.form.get('full_name')
        member.phone = request.form.get('phone')
        member.role = request.form.get('role', member.role)
        sup_id = request.form.get('supervisor_id')
        member.supervisor_id = int(sup_id) if sup_id else None
        member.is_active = request.form.get('is_active') == 'yes'

        # Permissions — checkbox tick hai to True
        for field, _label in PERMISSION_FIELDS:
            setattr(member, field, request.form.get(field) == 'yes')

        db.session.commit()
        log_activity('update', 'Team Member', member.full_name, 'Permissions / details update')
        flash(f'{member.full_name} ki details aur permissions update ho gayin', 'success')
        return redirect(url_for('team'))

    supervisors = User.query.filter_by(role='supervisor', is_active=True).all()
    return render_template('team_edit.html', member=member, supervisors=supervisors,
                           permission_fields=PERMISSION_FIELDS)


@app.route('/team/<int:member_id>/reset-password', methods=['POST'])
@login_required
@owner_required
def reset_member_password(member_id):
    """Owner kisi member ka password reset kar sakta hai"""
    member = User.query.get_or_404(member_id)
    if member.role == 'owner':
        flash('Owner ka password sirf Settings se change hota hai', 'error')
        return redirect(url_for('team'))

    new_pw = request.form.get('new_password') or ''
    if len(new_pw) < 6:
        flash('Password kam az kam 6 characters ka hona chahiye', 'error')
        return redirect(url_for('team'))

    member.set_password(new_pw)
    db.session.commit()
    log_activity('update', 'Team Member', member.full_name, 'Password reset kiya gaya')
    flash(f'{member.full_name} ka password reset ho gaya', 'success')
    return redirect(url_for('team'))


@app.route('/team/<int:member_id>/delete', methods=['POST'])
@login_required
@owner_required
def delete_team_member(member_id):
    """Team member delete — agar uske reports hon to sirf inactive karen"""
    member = User.query.get_or_404(member_id)
    if member.role == 'owner':
        flash('Owner account delete nahi ho sakta', 'error')
        return redirect(url_for('team'))

    report_count = DailyReport.query.filter_by(worker_id=member.id).count()
    page_count = Page.query.filter_by(assigned_worker_id=member.id).count()

    if report_count > 0 or page_count > 0:
        member.is_active = False
        db.session.commit()
        log_activity('update', 'Team Member', member.full_name,
                     f'Inactive kiya gaya ({report_count} reports, {page_count} pages linked)')
        flash(f'{member.full_name} ke {report_count} reports aur {page_count} pages linked hain — '
              f'delete ke bajaye INACTIVE kar diya gaya', 'warning')
        return redirect(url_for('team'))

    name = member.full_name
    db.session.delete(member)
    db.session.commit()
    log_activity('delete', 'Team Member', name)
    flash(f'{name} delete ho gaya', 'success')
    return redirect(url_for('team'))


# ==================== ACTIVITY LOG (OWNER ONLY) ====================
@app.route('/activity-log')
@login_required
@owner_required
def activity_log():
    """Kis ne kya kiya — sab record, khaas kar delete"""
    action_filter = request.args.get('action', '')
    query = ActivityLog.query
    if action_filter in ('create', 'update', 'delete'):
        query = query.filter_by(action=action_filter)

    logs = query.order_by(ActivityLog.created_at.desc()).limit(300).all()

    # Owner ne dekh liya — unseen mark hata dein
    unseen = ActivityLog.query.filter_by(seen_by_owner=False).all()
    for entry in unseen:
        entry.seen_by_owner = True
    if unseen:
        db.session.commit()

    counts = {
        'total': ActivityLog.query.count(),
        'delete': ActivityLog.query.filter_by(action='delete').count(),
        'create': ActivityLog.query.filter_by(action='create').count(),
        'update': ActivityLog.query.filter_by(action='update').count(),
    }
    return render_template('activity_log.html', logs=logs, counts=counts,
                           action_filter=action_filter)


# ==================== MASTER EXCEL EXPORT (OWNER ONLY) ====================
@app.route('/export/master')
@login_required
@owner_required
def export_master():
    """
    Ek hi Excel file mein SARA data — 9 sheets.
    Master sheet mein har page ke sath uski FB ID, BM, partner access,
    worker aur groups sab ek row mein.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    def add_sheet(title, headers, rows, widths, color='1877F2'):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths, 1):
            if i <= 26:
                ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ---------- SHEET 1: SUMMARY ----------
    all_accounts = FBAccount.query.all()
    all_pages = Page.query.all()
    all_bms = BusinessManager.query.all()
    all_partners = BMPartnerAccess.query.all()
    all_groups = Group.query.all()
    all_team = User.query.all()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(['FB MANAGER — COMPLETE DATA EXPORT'])
    ws_sum['A1'].font = Font(bold=True, size=16, color='1877F2')
    ws_sum.append([f'Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")}'])
    ws_sum.append([])

    summary_rows = [
        ('Total FB Accounts', len(all_accounts)),
        ('  — Active', sum(1 for a in all_accounts if a.status == 'active')),
        ('  — With Issues', sum(1 for a in all_accounts if a.issue_type and a.issue_type != 'none')),
        ('Total Pages', len(all_pages)),
        ('  — Monetized', sum(1 for p in all_pages if p.monetization_status == 'monetized')),
        ('  — Recommendation Not Okay', sum(1 for p in all_pages if (p.recommendation or 'okay') == 'not_okay')),
        ('  — Status Problem', sum(1 for p in all_pages if (p.page_status or 'active') != 'active')),
        ('Total Page Followers', sum(p.current_followers or 0 for p in all_pages)),
        ('Total Business Managers', len(all_bms)),
        ('  — Partner Access Records', len(all_partners)),
        ('  — BMs With Invites Given', sum(1 for b in all_bms if b.invited_to_fb_account_id)),
        ('Total Groups', len(all_groups)),
        ('  — Total Group Members', sum(g.members_count or 0 for g in all_groups)),
        ('Total Team Members', sum(1 for u in all_team if u.role != 'owner')),
        ('Total Investment (PKR)', sum(a.purchase_cost or 0 for a in all_accounts)),
    ]
    ws_sum.append(['Metric', 'Value'])
    for c in (1, 2):
        cell = ws_sum.cell(row=4, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1877F2', end_color='1877F2', fill_type='solid')
    for label, value in summary_rows:
        ws_sum.append([label, value])
    ws_sum.column_dimensions['A'].width = 32
    ws_sum.column_dimensions['B'].width = 20

    # ---------- SHEET 2: MASTER (page-centric, sab kuch jointly) ----------
    master_rows = []
    for p in all_pages:
        acc = p.fb_account
        bm = p.business_manager
        partner_names = ''
        if bm and bm.partner_accesses:
            partner_names = ', '.join(
                f"{pa.partner_bm_name}({pa.partner_bm_id or 'no-id'})" for pa in bm.partner_accesses
            )
        page_groups = [g for g in all_groups if g.page_id == p.id]
        group_txt = ', '.join(
            f"{g.group_name} ({g.members_count or 0:,}"
            + (f", {len(g.all_account_ids)} IDs" if len(g.all_account_ids) > 1 else "") + ")"
            for g in page_groups
        )
        pending_emails = ', '.join(i.email for i in bm.pending_invites) if bm else ''

        master_rows.append([
            p.page_name, p.page_url or '', p.niche or '',
            p.page_status or 'active', p.recommendation or 'okay',
            p.monetization_status or 'non_monetized',
            'Yes' if p.is_fresh_start else 'No',
            p.followers_at_start or 0, p.current_followers or 0,
            # FB ID
            acc.account_name if acc else '', acc.email if acc else '',
            acc.location if acc else '', acc.profile_username if acc else '',
            (acc.issue_type if acc and acc.issue_type != 'none' else '') if acc else '',
            acc.status if acc else '',
            # BM
            bm.bm_name if bm else '', bm.bm_id if bm else '',
            bm.invited_fb_account.account_name if (bm and bm.invited_fb_account) else '',
            pending_emails,
            len(bm.partner_accesses) if bm else 0, partner_names,
            # worker + groups
            p.assigned_worker.full_name if p.assigned_worker else '',
            len(page_groups), group_txt,
            p.page_created_date.strftime('%Y-%m-%d') if p.page_created_date else '',
            p.monetized_date.strftime('%Y-%m-%d') if p.monetized_date else '',
            p.notes or ''
        ])

    add_sheet("Master Sheet", [
        'Page Name', 'Page URL', 'Niche', 'Page Status', 'Recommendation', 'Monetization',
        'Fresh Start', 'Followers At Start', 'Current Followers',
        'FB ID Name', 'FB Email', 'FB Location', 'FB Username', 'FB Issue', 'FB Status',
        'BM Name', 'BM ID', 'BM Invited To (FB ID)', 'BM Pending Invite Emails',
        'Partner Access Count', 'Partner BMs',
        'Assigned Worker', 'Groups Count', 'Linked Groups', 'Page Created', 'Monetized Date', 'Notes'
    ], master_rows,
        [26, 32, 14, 14, 16, 16, 12, 16, 16, 22, 24, 14, 20, 16, 12,
         22, 20, 22, 34, 16, 40, 20, 12, 40, 14, 14, 30], '1877F2')

    # ---------- SHEET 3: FB ACCOUNTS ----------
    add_sheet("FB Accounts", [
        'Account Name', 'Email', 'Phone', 'Password', 'Recovery Email', '2FA Code',
        'Location', 'Issue Type', 'Issue Notes', 'Profile Link', 'Username', 'DOB',
        'Status', 'Purchase Date', 'Cost (PKR)', 'Pages', 'BMs', 'Notes'
    ], [[
        a.account_name or '', a.email or '', a.phone or '', a.get_fb_password() or '',
        a.recovery_email or '', a.get_two_fa_code() or '', a.location or '',
        a.issue_type or 'none', a.issue_notes or '', a.profile_link or '',
        a.profile_username or '', a.date_of_birth.strftime('%Y-%m-%d') if a.date_of_birth else '',
        a.status or 'active', a.purchase_date.strftime('%Y-%m-%d') if a.purchase_date else '',
        a.purchase_cost or 0, len(a.pages), len(a.business_managers), a.notes or ''
    ] for a in all_accounts],
        [22, 26, 15, 20, 24, 20, 14, 16, 28, 34, 18, 12, 12, 14, 14, 8, 8, 28], 'd62976')

    # ---------- SHEET 4: BUSINESS MANAGERS ----------
    add_sheet("Business Managers", [
        'BM Name', 'BM ID', 'Owner FB ID', 'Status', 'Invited To FB ID', 'Invite Date',
        'Invite Notes', 'Partner Access Count', 'Pages Under BM'
    ], [[
        b.bm_name or '', b.bm_id or '',
        b.fb_account.account_name if b.fb_account else '', b.status or '',
        b.invited_fb_account.account_name if b.invited_fb_account else '',
        b.invite_date.strftime('%Y-%m-%d') if b.invite_date else '',
        b.invite_notes or '', len(b.partner_accesses),
        Page.query.filter_by(bm_id=b.id).count()
    ] for b in all_bms],
        [24, 22, 22, 12, 24, 14, 30, 18, 16], '7F77DD')

    # ---------- SHEET 5: PARTNER ACCESS ----------
    add_sheet("Partner Access", [
        'Source BM Name', 'Source BM ID', 'Source Owner FB ID', 'Partner BM Name',
        'Partner BM ID', 'Access Level', 'Granted Date', 'Active', 'Notes'
    ], [[
        pa.source_bm.bm_name if pa.source_bm else '',
        pa.source_bm.bm_id if pa.source_bm else '',
        pa.source_bm.fb_account.account_name if (pa.source_bm and pa.source_bm.fb_account) else '',
        pa.partner_bm_name or '', pa.partner_bm_id or '', pa.access_level or '',
        pa.access_granted_date.strftime('%Y-%m-%d') if pa.access_granted_date else '',
        'Yes' if pa.is_active else 'No', pa.notes or ''
    ] for pa in all_partners],
        [24, 20, 22, 24, 20, 18, 14, 10, 30], '5b21b6')

    # ---------- SHEET 6: PAGES ----------
    add_sheet("Pages", [
        'Page Name', 'Page URL', 'Niche', 'FB ID', 'BM', 'Worker', 'Monetization',
        'Page Status', 'Status Notes', 'Recommendation', 'Reco Notes',
        'Fresh Start', 'Followers Start', 'Current Followers', 'Created', 'Monetized', 'Notes'
    ], [[
        p.page_name or '', p.page_url or '', p.niche or '',
        p.fb_account.account_name if p.fb_account else '',
        p.business_manager.bm_name if p.business_manager else '',
        p.assigned_worker.full_name if p.assigned_worker else '',
        p.monetization_status or '', p.page_status or 'active', p.page_status_notes or '',
        p.recommendation or 'okay', p.recommendation_notes or '',
        'Yes' if p.is_fresh_start else 'No', p.followers_at_start or 0, p.current_followers or 0,
        p.page_created_date.strftime('%Y-%m-%d') if p.page_created_date else '',
        p.monetized_date.strftime('%Y-%m-%d') if p.monetized_date else '', p.notes or ''
    ] for p in all_pages],
        [26, 32, 14, 22, 20, 18, 16, 14, 26, 16, 26, 12, 15, 16, 14, 14, 28], '42b72a')

    # ---------- SHEET 7: GROUPS ----------
    def _group_all_ids(g):
        names = []
        if g.fb_account:
            names.append(g.fb_account.account_name + ' (main)')
        for ga in g.extra_accounts:
            if ga.fb_account:
                names.append(ga.fb_account.account_name)
        return ', '.join(names)

    add_sheet("Groups", [
        'Group Name', 'Group URL', 'Group FB ID', 'Linked Page', 'Main FB ID',
        'Total FB IDs', 'All FB IDs (admin)', 'Members', 'Status', 'Notes'
    ], [[
        g.group_name or '', g.group_url or '', g.group_fb_id or '',
        g.page.page_name if g.page else '', g.fb_account.account_name if g.fb_account else '',
        len(g.all_account_ids), _group_all_ids(g),
        g.members_count or 0, g.status or 'active', g.notes or ''
    ] for g in all_groups],
        [28, 34, 20, 24, 22, 12, 45, 14, 12, 30], '0ea5e9')

    # ---------- SHEET: BM EMAIL INVITES ----------
    all_invites = BMInvite.query.all()
    add_sheet("BM Email Invites", [
        'BM Name', 'BM ID', 'Invite Email', 'Status', 'Invited Date', 'Accepted Date',
        'Linked FB ID', 'Notes'
    ], [[
        inv.business_manager.bm_name if inv.business_manager else '',
        inv.business_manager.bm_id if inv.business_manager else '',
        inv.email or '', (inv.status or 'pending').title(),
        inv.invited_date.strftime('%Y-%m-%d') if inv.invited_date else '',
        inv.accepted_date.strftime('%Y-%m-%d') if inv.accepted_date else '',
        inv.fb_account.account_name if inv.fb_account else '',
        inv.notes or ''
    ] for inv in all_invites],
        [24, 20, 34, 14, 14, 14, 22, 30], '16a34a')

    # ---------- SHEET 8: TEAM ----------
    add_sheet("Team", [
        'Full Name', 'Username', 'Role', 'Phone', 'Supervisor', 'Active', 'Joined',
        'Pages Assigned', 'Can Add/Edit', 'Can Delete', 'Can Export', 'Can View Passwords'
    ], [[
        u.full_name or '', u.username or '', u.role or '', u.phone or '',
        u.supervisor.full_name if u.supervisor else '', 'Yes' if u.is_active else 'No',
        u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
        Page.query.filter_by(assigned_worker_id=u.id).count(),
        'Yes' if u.has_perm('add_edit') else 'No',
        'Yes' if u.has_perm('delete') else 'No',
        'Yes' if u.has_perm('export') else 'No',
        'Yes' if u.has_perm('view_passwords') else 'No'
    ] for u in all_team],
        [24, 18, 14, 15, 22, 10, 14, 14, 14, 12, 12, 18], 'ec4899')

    # ---------- SHEET 9: PAYMENTS ----------
    add_sheet("Payments", [
        'Month', 'Member', 'Role', 'Earned (USD)', 'To Pay (PKR)', 'FB Received',
        'Received Date', 'Paid', 'Payment Date', 'Method', 'Reference', 'Status', 'Notes'
    ], [[
        pm.month or '', pm.user.full_name if pm.user else '', pm.user.role if pm.user else '',
        pm.total_earned_usd or 0, pm.agreed_amount_pkr or 0,
        'Yes' if pm.received_from_fb else 'No',
        pm.received_date.strftime('%Y-%m-%d') if pm.received_date else '',
        'Yes' if pm.paid_to_member else 'No',
        pm.payment_date.strftime('%Y-%m-%d') if pm.payment_date else '',
        pm.payment_method or '', pm.payment_reference or '', pm.status, pm.notes or ''
    ] for pm in TeamPayment.query.all()],
        [12, 22, 14, 14, 16, 12, 14, 10, 14, 16, 20, 12, 26], 'f59e0b')

    # ---------- SHEET 10: DAILY REPORTS (last 90 days) ----------
    since = date.today() - timedelta(days=90)
    reports = DailyReport.query.filter(DailyReport.report_date >= since).order_by(
        DailyReport.report_date.desc()).all()
    add_sheet("Daily Reports (90d)", [
        'Date', 'Page', 'FB ID', 'Worker', 'Views', 'Reach', 'Followers Gained', 'Earnings (USD)', 'Notes'
    ], [[
        r.report_date.strftime('%Y-%m-%d') if r.report_date else '',
        r.page.page_name if r.page else '',
        r.page.fb_account.account_name if (r.page and r.page.fb_account) else '',
        r.worker.full_name if r.worker else '',
        r.views or 0, r.reach or 0, r.followers_gained or 0, r.earnings_usd or 0, r.notes or ''
    ] for r in reports],
        [14, 26, 22, 20, 12, 12, 16, 15, 26], '0891b2')

    # ---------- SHEET 11: ACTIVITY LOG ----------
    add_sheet("Activity Log", [
        'Date & Time', 'User', 'Role', 'Action', 'Type', 'Name', 'Details'
    ], [[
        l.created_at.strftime('%Y-%m-%d %H:%M') if l.created_at else '',
        l.user_name or '', l.user_role or '', (l.action or '').upper(),
        l.entity_type or '', l.entity_name or '', l.details or ''
    ] for l in ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(1000).all()],
        [18, 22, 14, 12, 18, 30, 40], '64748b')

    filename = f"FB_Manager_COMPLETE_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    log_activity('create', 'Master Export', filename, 'Complete data export download kiya')
    return _excel_response(wb, filename)


# ==================== INITIALIZE ====================
@app.cli.command('init-db')
def init_db_command():
    """Database initialize karne ke liye"""
    init_db(app)


with app.app_context():
    db.create_all()
    init_db(app)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
